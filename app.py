import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, Scrollbar
from PIL import Image as pilimage, ImageTk
import re
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import pandas as pd
import sys
import os

def caminho_recurso(relativo):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relativo)
    return os.path.join(os.path.abspath("."), relativo)
######################################################################### EXTRAS ###########################################################################################################################
# simbolo procurar
img_procurar = pilimage.open(caminho_recurso('imagens/procurar.webp'))
simbolo_procurar = ctk.CTkImage(img_procurar, size=(20,20))
# simbolo adicionar 
img_adicionar = pilimage.open(caminho_recurso('imagens/adicionar.png'))
simbolo_adicionar = ctk.CTkImage(img_adicionar, size=(20,20))
# simbolo confirmar
img_confirmar = pilimage.open(caminho_recurso('imagens/confirmar.png'))
simbolo_confirmar = ctk.CTkImage(img_confirmar, size=(10,10))
# simbolo cancelar
img_cancelar = pilimage.open(caminho_recurso('imagens/cancelar.png'))
simbolo_cancelar = ctk.CTkImage(img_cancelar, size=(10,10))
# simbolo calcular
img_calcular = pilimage.open(caminho_recurso('imagens/calculadora.png'))
simbolo_calcular = ctk.CTkImage(img_calcular, size=(15,15))
# tematica 
cor_sistema= 'dark'


#BANCOS DE DADOS 
# MOTORISTAS
banco_mot = sqlite3.connect(caminho_recurso('MOTORISTAS.db'))
cursor_mot = banco_mot.cursor()
cursor_mot.execute('''CREATE TABLE IF NOT EXISTS MOTORISTAS (
"CÓDIGO" TEXT, 
"DATA DO CADASTRO" TEXT, 
"NOME" TEXT,
"CPF" TEXT,
"IDENTIDADE" TEXT,
"DATA DO NASCIMENTO" TEXT,
"TELEFONE 1" TEXT,
"TELEFONE 2" TEXT,
"TELEFONE 3" TEXT,
"EMAIL" TEXT,
"CIDADE" TEXT,
"BAIRRO" TEXT, 
"RUA" TEXT,
"Nº DA CASA" TEXT,
"CEP" TEXT,
"INFORMAÇÕES ADICIONAIS" TEXT
)
''')

# PROPRIETARIOS
banco_prop = sqlite3.connect(caminho_recurso('PROPRIETARIOS.db'))
cursor_prop = banco_prop.cursor()
cursor_prop.execute('''CREATE TABLE IF NOT EXISTS PROPRIETARIOS(
"CÓDIGO" TEXT, 
"DATA DO CADASTRO" TEXT, 
"NOME" TEXT,
"CPF/CNPJ" TEXT,
"IDENTIDADE" TEXT,
"DATA DO NASCIMENTO" TEXT,
"TELEFONE 1" TEXT,
"TELEFONE 2" TEXT,
"TELEFONE 3" TEXT,
"EMAIL" TEXT,
"CIDADE" TEXT,
"BAIRRO" TEXT, 
"RUA" TEXT,
"Nº DA CASA" TEXT,
"CEP" TEXT,
"INFORMAÇÕES ADICIONAIS" TEXT
)
''')

# VEICULOS
banco_vei = sqlite3.connect(caminho_recurso('VEICULOS.db'))
cursor_vei = banco_vei.cursor()
cursor_vei.execute('''CREATE TABLE IF NOT EXISTS VEICULOS (
"CÓDIGO" TEXT, 
"MARCA" TEXT, 
"MODELO" TEXT,
"ANO" TEXT,
"PLACA" TEXT,
"CHASSI" TEXT,
"RENAVAN" TEXT
)
''')

# CLIENTES
banco_cli = sqlite3.connect(caminho_recurso('CLIENTES.db'))
cursor_cli = banco_cli.cursor()
cursor_cli.execute('''CREATE TABLE IF NOT EXISTS CLIENTES (
"CÓDIGO" TEXT,
"NOME" TEXT, 
"CNPJ" TEXT, 
"INSC. ESTADUAL" TEXT,
"ENDEREÇO" TEXT,
"MUNICIPIO" TEXT
)
''')

# numero do recibo

banco_recibo =  sqlite3.connect(caminho_recurso('NUMERO_RECIBO.db'))
cursor_recibo = banco_recibo.cursor()
cursor_recibo.execute('CREATE TABLE IF NOT EXISTS NUMERO_RECIBO ("NUMERO DO RECIBO" TEXT)')

def gerar_codigo_motorista():
    cursor_mot.execute('SELECT MAX("CÓDIGO") FROM MOTORISTAS')
    resultado = cursor_mot.fetchone()[0]
    if resultado is None:
        return '0001'
    return f'{int(resultado)+1:04}'

def gerar_codigo_proprietario():
    cursor_prop.execute('SELECT MAX("CÓDIGO") FROM PROPRIETARIOS')
    resultado = cursor_prop.fetchone()[0]
    if resultado is None:
        return '0001'
    return f'{int(resultado)+1:04}'
def gerar_codigo_veiculo():
    cursor_vei.execute('SELECT MAX("CÓDIGO") FROM VEICULOS')
    resultado = cursor_vei.fetchone()[0]
    if resultado is None:
        return '0001'
    return f'{int(resultado)+1:04}'

def gerar_codigo_cliente():
    cursor_cli.execute('SELECT MAX("CÓDIGO") FROM CLIENTES')
    resultado = cursor_cli.fetchone()[0]
    if resultado is None:
        return '0001'
    return f'{int(resultado)+1:04}'

def gerar_numero_recibo():
    cursor_recibo.execute('SELECT MAX("NUMERO DO RECIBO") FROM NUMERO_RECIBO')
    resultado = cursor_recibo.fetchone()[0]
    if resultado is None:
        return '0001'
    return f'{int(resultado)+1:04}'


    
######################################################################### EMISSÃO DE CARTA FRETE #################################################################################################################


# definir tela principal
app_2 = ctk.CTk()

app_2._set_appearance_mode(cor_sistema)

app_2.title('CADASTRO DE MOTORISTAS')

# atribuir tamanho a tela principal
largura_app_2 = 1600
altura_app_2 = 900

# infirmações do tamanho da tela
largura_tela_2 = app_2.winfo_screenwidth()
altura_tela_2 = app_2.winfo_screenheight()

# centralizar tela principal
x_2 = int((largura_tela_2 - largura_app_2) / 2)
y_2 = int((altura_tela_2 - altura_app_2) / 2)

# definir tamanho ja centralizado da tela principal
app_2.geometry(f'{largura_app_2}x{altura_app_2}+{x_2}+{y_2}')
app_2.minsize(width= 1600, height=900)

logo_img = pilimage.open(caminho_recurso('imagens/logo.PNG'))
logo = ctk.CTkImage(logo_img, size=(200,50))
logo_ = ctk.CTkLabel(app_2, image=logo, text='')
logo_.place(x=40,y=20)

######################### Dados do motorista ###############################


# titulo block dados do motorista
titulo_frame = ctk.CTkLabel(master= app_2, text='Dados do Motorista', font=('Arial', 14,'bold') )
titulo_frame.pack(pady= (10,0))

# frame de dados do motorista 
dados_motorista = ctk.CTkFrame(app_2, width=1050, height=250, border_width=2)
dados_motorista.pack(pady= 10)

# codigo do motorista
texto_codigo_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Código' )
texto_codigo_motorista.place(x= 43, y= 20)
container_codigo_motorista = ctk.CTkEntry(master= dados_motorista, width=50,)
container_codigo_motorista.place(x= 40, y= 45)

# inserir data do cadastro do motorista
def formatar_data(event=None):
    texto = container_data_cadastro_motorista.get()
    texto_limpo = ''.join(filter(str.isdigit, texto))[:8]  # Máximo 8 dígitos

    texto_formatado = ''
    if len(texto_limpo) >= 1:
        texto_formatado += texto_limpo[:2]
    if len(texto_limpo) >= 3:
        texto_formatado = texto_formatado[:2] + '/' + texto_limpo[2:4]
    if len(texto_limpo) >= 5:
        texto_formatado = texto_formatado[:5] + '/' + texto_limpo[4:8]

    container_data_cadastro_motorista.delete(0, 'end')
    container_data_cadastro_motorista.insert(0, texto_formatado)
texto_data_cadastro_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Data do Cadastro' )
texto_data_cadastro_motorista.place(x= 103, y= 20)
container_data_cadastro_motorista = ctk.CTkEntry(master= dados_motorista)
container_data_cadastro_motorista.place(x= 100, y= 45)
container_data_cadastro_motorista.bind("<KeyRelease>", formatar_data)

# inserir nome do motorista
texto_nome_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Nome do motorista' )
texto_nome_motorista.place(x= 253, y= 20)
container_nome_motorista = ctk.CTkEntry(master= dados_motorista)
container_nome_motorista.place(x= 250, y= 45)

# inserir CPF do motorista
def formatar_cpf_motorista(event=None):
    texto = container_cpf_motorista.get().replace(".", "").replace("-", "")
    texto = re.sub(r'\D', '', texto)

    novo_texto = ''
    if len(texto) > 0:
        novo_texto += texto[:3]
    if len(texto) >= 4:
        novo_texto = novo_texto[:3] + '.' + texto[3:6]
    if len(texto) >= 7:
        novo_texto = novo_texto[:7] + '.' + texto[6:9]
    if len(texto) >= 10:
        novo_texto = novo_texto[:11] + '-' + texto[9:11]

    container_cpf_motorista.delete(0, 'end')
    container_cpf_motorista.insert(0, novo_texto)

texto_cpf_motorista = ctk.CTkLabel(master= dados_motorista, text= 'CPF')
texto_cpf_motorista.place(x= 403, y= 20)
container_cpf_motorista = ctk.CTkEntry(master= dados_motorista)
container_cpf_motorista.place(x= 400, y= 45)

# Chama a função toda vez que o usuário digita algo
container_cpf_motorista.bind("<KeyRelease>", formatar_cpf_motorista)

# inserir Identidade do motorista
def formatar_identidade_motorista(event=None):
    texto = container_identidade_motorista.get()

    # Remove tudo que não for número
    texto = re.sub(r'\D', '', texto)
    container_identidade_motorista.delete(0, 'end')
    container_identidade_motorista.insert(0, texto)
    
texto_identidade_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Identidade' )
texto_identidade_motorista.place(x= 553, y= 20)
container_identidade_motorista = ctk.CTkEntry(master= dados_motorista)
container_identidade_motorista.place(x= 550, y= 45)
container_identidade_motorista.bind("<KeyRelease>", formatar_identidade_motorista)


# inserir data do nascimento do motorista
def formatar_data_nascimento(event=None):
    texto = container_data_nascimento_motorista.get()
    texto_limpo = ''.join(filter(str.isdigit, texto))[:8]  # Máximo 8 dígitos

    texto_formatado = ''
    if len(texto_limpo) >= 1:
        texto_formatado += texto_limpo[:2]
    if len(texto_limpo) >= 3:
        texto_formatado = texto_formatado[:2] + '/' + texto_limpo[2:4]
    if len(texto_limpo) >= 5:
        texto_formatado = texto_formatado[:5] + '/' + texto_limpo[4:8]

    container_data_nascimento_motorista.delete(0, 'end')
    container_data_nascimento_motorista.insert(0, texto_formatado)
texto_data_nascimento_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Data do nascimento' )
texto_data_nascimento_motorista.place(x= 703, y= 20)
container_data_nascimento_motorista = ctk.CTkEntry(master= dados_motorista)
container_data_nascimento_motorista.place(x= 700, y= 45)
container_data_nascimento_motorista.bind("<KeyRelease>", formatar_data_nascimento)

# inserir Telefone do motorista

def formatar_telefone_motorista(event=None):
    texto = container_telefone_motorista.get()

    # Remove tudo que não for número
    texto = re.sub(r'\D', '', texto)

    novo_texto = ''
    if len(texto) >= 1:
        novo_texto = '(' + texto[:2]  # DDD
    if len(texto) >= 3:
        if len(texto) >= 11:  # celular com 9 dígitos
            novo_texto += ') ' + texto[2:7] + '-' + texto[7:11]
        elif len(texto) >= 10:  # fixo com 8 dígitos
            novo_texto += ') ' + texto[2:6] + '-' + texto[6:10]
        elif len(texto) > 2:  # ainda incompleto, mas já dá pra exibir parcialmente
            novo_texto += ') ' + texto[2:]

    container_telefone_motorista.delete(0, 'end')
    container_telefone_motorista.insert(0, novo_texto)
texto_telefone_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Telefone' )
texto_telefone_motorista.place(x= 853, y= 20)
container_telefone_motorista = ctk.CTkEntry(master= dados_motorista)
container_telefone_motorista.place(x= 850, y= 45)
container_telefone_motorista.bind("<KeyRelease>", formatar_telefone_motorista)

# inserir Telefone 2 motorista
def formatar_telefone_2_motorista(event=None):
    texto = container_telefone_2_motorista.get()

    # Remove tudo que não for número
    texto = re.sub(r'\D', '', texto)

    novo_texto = ''
    if len(texto) >= 1:
        novo_texto = '(' + texto[:2]  # DDD
    if len(texto) >= 3:
        if len(texto) >= 11:  # celular com 9 dígitos
            novo_texto += ') ' + texto[2:7] + '-' + texto[7:11]
        elif len(texto) >= 10:  # fixo com 8 dígitos
            novo_texto += ') ' + texto[2:6] + '-' + texto[6:10]
        elif len(texto) > 2:  # ainda incompleto, mas já dá pra exibir parcialmente
            novo_texto += ') ' + texto[2:]

    container_telefone_2_motorista.delete(0, 'end')
    container_telefone_2_motorista.insert(0, novo_texto)
texto_telefone_2_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Telefone 2' )
texto_telefone_2_motorista.place(x= 43, y= 90)
container_telefone_2_motorista = ctk.CTkEntry(master= dados_motorista)
container_telefone_2_motorista.place(x= 40, y= 115)
container_telefone_2_motorista.bind("<KeyRelease>", formatar_telefone_2_motorista)

# inserir Telefone 3 motorista
def formatar_telefone_3_motorista(event=None):
    texto = container_telefone_3_motorista.get()

    # Remove tudo que não for número
    texto = re.sub(r'\D', '', texto)

    novo_texto = ''
    if len(texto) >= 1:
        novo_texto = '(' + texto[:2]  # DDD
    if len(texto) >= 3:
        if len(texto) >= 11:  # celular com 9 dígitos
            novo_texto += ') ' + texto[2:7] + '-' + texto[7:11]
        elif len(texto) >= 10:  # fixo com 8 dígitos
            novo_texto += ') ' + texto[2:6] + '-' + texto[6:10]
        elif len(texto) > 2:  # ainda incompleto, mas já dá pra exibir parcialmente
            novo_texto += ') ' + texto[2:]

    container_telefone_3_motorista.delete(0, 'end')
    container_telefone_3_motorista.insert(0, novo_texto)
texto_telefone_3_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Telefone 3' )
texto_telefone_3_motorista.place(x= 193, y= 90)
container_telefone_3_motorista = ctk.CTkEntry(master= dados_motorista)
container_telefone_3_motorista.place(x= 190, y= 115)
container_telefone_3_motorista.bind("<KeyRelease>", formatar_telefone_3_motorista)

# inserir Email do motorista
texto_email_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Email' )
texto_email_motorista.place(x= 343, y= 90)
container_email_motorista = ctk.CTkEntry(master= dados_motorista)
container_email_motorista.place(x= 340, y= 115)

# inserir cidade do motorista
texto_cidade_motorista = ctk.CTkLabel(master=dados_motorista, text='Cidade')
texto_cidade_motorista.place(x= 493, y= 90)
container_cidade_motorista = ctk.CTkEntry(master= dados_motorista)
container_cidade_motorista.place(x= 490, y= 115)

# inserir bairro do motorista
texto_bairro_motorista = ctk.CTkLabel(master=dados_motorista, text='Bairro')
texto_bairro_motorista.place(x= 643, y= 90)
container_bairro_motorista = ctk.CTkEntry(master= dados_motorista)
container_bairro_motorista.place(x= 640, y= 115)


# inserir rua da casa do motorista
texto_rua_casa_motorista = ctk.CTkLabel(master=dados_motorista, text='Rua')
texto_rua_casa_motorista.place(x= 793, y= 90)
container_rua_casa_motorista = ctk.CTkEntry(master= dados_motorista)
container_rua_casa_motorista.place(x= 790, y= 115)

# inserir numero da casa do motorista
def formatar_numero_casa_motorista(event=None):
    texto = container_numero_casa_motorista.get()

    # Remove tudo que não for número
    texto = re.sub(r'\D', '', texto)
    container_numero_casa_motorista.delete(0, 'end')
    container_numero_casa_motorista.insert(0, texto)
    
texto_numero_casa_motorista = ctk.CTkLabel(master=dados_motorista, text='Nº Casa')
texto_numero_casa_motorista.place(x= 43, y= 160)
container_numero_casa_motorista = ctk.CTkEntry(master= dados_motorista)
container_numero_casa_motorista.place(x= 40, y= 185)
container_numero_casa_motorista.bind("<KeyRelease>", formatar_numero_casa_motorista)

# inserir cep motorista
def formatar_cep_motorista(event=None):
    texto = container_cep_motorista.get()

    # Remove tudo que não for número
    texto = re.sub(r'\D', '', texto)
    container_cep_motorista.delete(0, 'end')
    container_cep_motorista.insert(0, texto)
    
texto_cep_motorista = ctk.CTkLabel(master=dados_motorista, text='CEP')
texto_cep_motorista.place(x= 193, y= 160)
container_cep_motorista = ctk.CTkEntry(master= dados_motorista)
container_cep_motorista.place(x= 190, y= 185)
container_cep_motorista.bind("<KeyRelease>", formatar_cep_motorista)

# inserir informações adicionais motorista 
texto_informacoes_adicionais_motorista = ctk.CTkLabel(master=dados_motorista, text='Informações adicionais')
texto_informacoes_adicionais_motorista.place(x= 343, y= 160)
container_informacoes_adicionais_motorista = ctk.CTkTextbox(master= dados_motorista, height= 50, width=140, border_width=2)
container_informacoes_adicionais_motorista.place(x= 340, y= 185)

def limpar_conteudo():
    container_codigo_motorista.configure(state='normal')
    container_codigo_motorista.delete(0, 'end')
    container_data_cadastro_motorista.configure(state='normal')
    container_data_cadastro_motorista.delete(0, 'end')
    container_nome_motorista.configure(state='normal')
    container_nome_motorista.delete(0, 'end')
    container_cpf_motorista.configure(state='normal')
    container_cpf_motorista.delete(0, 'end')
    container_identidade_motorista.configure(state='normal')
    container_identidade_motorista.delete(0, 'end')
    container_data_nascimento_motorista.configure(state='normal')
    container_data_nascimento_motorista.delete(0, 'end')
    container_telefone_motorista.configure(state='normal')
    container_telefone_motorista.delete(0, 'end')
    container_telefone_2_motorista.configure(state='normal')
    container_telefone_2_motorista.delete(0, 'end')
    container_telefone_3_motorista.configure(state='normal')
    container_telefone_3_motorista.delete(0, 'end')
    container_email_motorista.configure(state='normal')
    container_email_motorista.delete(0, 'end')
    container_cidade_motorista.configure(state='normal')
    container_cidade_motorista.delete(0, 'end')
    container_bairro_motorista.configure(state='normal')
    container_bairro_motorista.delete(0, 'end')
    container_rua_casa_motorista.configure(state='normal')
    container_rua_casa_motorista.delete(0, 'end')
    container_numero_casa_motorista.configure(state='normal')
    container_numero_casa_motorista.delete(0, 'end')
    container_cep_motorista.configure(state='normal')
    container_cep_motorista.delete(0, 'end')
    container_informacoes_adicionais_motorista.configure(state='normal')
    container_informacoes_adicionais_motorista.delete("1.0", "end")     
    
botao_limpar = ctk.CTkButton(dados_motorista, text='Limpar', fg_color='light yellow', font=('Arial', 10), text_color='black', width=30, hover_color='grey', command=limpar_conteudo)
botao_limpar.place(x=1000, y=4)
    


############################################################### BOTAO CADASTRAR MOTORISTA #######################################################################################################################
def cadastrar_motorista():
    
    app_3 = ctk.CTkToplevel(app_2)
    app_3.lift()
    app_3.focus_force()
    app_3.grab_set()
    app_3.title('Cadastrar Motorista')
    app_3.maxsize(width=1000, height=300)
    # frame de dados do motorista 
    dados_motorista = ctk.CTkFrame(app_3, width=1050, height=250, border_width=2)
    dados_motorista.pack(pady= 10)
    codigo_motorista = gerar_codigo_motorista()
    
    # codigo do motorista
    texto_codigo_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Código' )
    texto_codigo_motorista.place(x= 43, y= 20)
    container_codigo_motorista = ctk.CTkEntry(master= dados_motorista, width=50)
    container_codigo_motorista.insert(0, codigo_motorista)
    container_codigo_motorista.configure(state='readonly', text_color= 'grey')
    container_codigo_motorista.place(x= 40, y= 45)


    # inserir data do cadastro do motorista
    def formatar_data(event=None):
        texto = container_data_cadastro_motorista.get()
        texto_limpo = ''.join(filter(str.isdigit, texto))[:8]  # Máximo 8 dígitos

        texto_formatado = ''
        if len(texto_limpo) >= 1:
            texto_formatado += texto_limpo[:2]
        if len(texto_limpo) >= 3:
            texto_formatado = texto_formatado[:2] + '/' + texto_limpo[2:4]
        if len(texto_limpo) >= 5:
            texto_formatado = texto_formatado[:5] + '/' + texto_limpo[4:8]

        container_data_cadastro_motorista.delete(0, 'end')
        container_data_cadastro_motorista.insert(0, texto_formatado)
        
    texto_data_cadastro_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Data do Cadastro' )
    texto_data_cadastro_motorista.place(x= 103, y= 20)
    container_data_cadastro_motorista = ctk.CTkEntry(master= dados_motorista, fg_color='white', text_color='black')
    container_data_cadastro_motorista.place(x= 100, y= 45)
    container_data_cadastro_motorista.bind("<KeyRelease>", formatar_data)
    

    # inserir nome do motorista
    texto_nome_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Nome do Motorista')
    texto_nome_motorista.place(x= 253, y= 20)
    container_nome_motorista = ctk.CTkEntry(master= dados_motorista, fg_color='white', text_color='black')
    container_nome_motorista.place(x= 250, y= 45)
    

   # inserir CPF do motorista
    def formatar_cpf_motorista(event=None):
        texto = container_cpf_motorista.get().replace(".", "").replace("-", "")
        texto = re.sub(r'\D', '', texto)

        novo_texto = ''
        if len(texto) > 0:
            novo_texto += texto[:3]
        if len(texto) >= 4:
            novo_texto = novo_texto[:3] + '.' + texto[3:6]
        if len(texto) >= 7:
            novo_texto = novo_texto[:7] + '.' + texto[6:9]
        if len(texto) >= 10:
            novo_texto = novo_texto[:11] + '-' + texto[9:11]

        container_cpf_motorista.delete(0, 'end')
        container_cpf_motorista.insert(0, novo_texto)

    texto_cpf_motorista = ctk.CTkLabel(master= dados_motorista, text= 'CPF')
    texto_cpf_motorista.place(x= 403, y= 20)
    container_cpf_motorista = ctk.CTkEntry(master= dados_motorista, fg_color='white', text_color='black')
    container_cpf_motorista.place(x= 400, y= 45)
    container_cpf_motorista.bind("<KeyRelease>", formatar_cpf_motorista)
    

    # inserir Identidade do motorista
    def formatar_identidade_motorista(event=None):
        texto = container_identidade_motorista.get()

        # Remove tudo que não for número
        texto = re.sub(r'\D', '', texto)
        container_identidade_motorista.delete(0, 'end')
        container_identidade_motorista.insert(0, texto)
        
    texto_identidade_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Identidade' )
    texto_identidade_motorista.place(x= 553, y= 20)
    container_identidade_motorista = ctk.CTkEntry(master= dados_motorista, fg_color='white', text_color='black')
    container_identidade_motorista.place(x= 550, y= 45)
    container_identidade_motorista.bind('<KeyRelease>', formatar_identidade_motorista)
   

    # inserir data do nascimento do motorista
    def formatar_data_nascimento(event=None):
        texto = container_data_nascimento_motorista.get()
        texto_limpo = ''.join(filter(str.isdigit, texto))[:8]  # Máximo 8 dígitos

        texto_formatado = ''
        if len(texto_limpo) >= 1:
            texto_formatado += texto_limpo[:2]
        if len(texto_limpo) >= 3:
            texto_formatado = texto_formatado[:2] + '/' + texto_limpo[2:4]
        if len(texto_limpo) >= 5:
            texto_formatado = texto_formatado[:5] + '/' + texto_limpo[4:8]

        container_data_nascimento_motorista.delete(0, 'end')
        container_data_nascimento_motorista.insert(0, texto_formatado)
    texto_data_nascimento_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Data do nascimento' )
    texto_data_nascimento_motorista.place(x= 703, y= 20)
    container_data_nascimento_motorista = ctk.CTkEntry(master= dados_motorista, fg_color='white', text_color='black')
    container_data_nascimento_motorista.place(x= 700, y= 45)
    container_data_nascimento_motorista.bind("<KeyRelease>", formatar_data_nascimento)
    
    
    # inserir Telefone do motorista

    def formatar_telefone_motorista(event=None):
        texto = container_telefone_motorista.get()

        # Remove tudo que não for número
        texto = re.sub(r'\D', '', texto)

        novo_texto = ''
        if len(texto) >= 1:
            novo_texto = '(' + texto[:2]  # DDD
        if len(texto) >= 3:
            if len(texto) >= 11:  # celular com 9 dígitos
                novo_texto += ') ' + texto[2:7] + '-' + texto[7:11]
            elif len(texto) >= 10:  # fixo com 8 dígitos
                novo_texto += ') ' + texto[2:6] + '-' + texto[6:10]
            elif len(texto) > 2:  # ainda incompleto, mas já dá pra exibir parcialmente
                novo_texto += ') ' + texto[2:]

        container_telefone_motorista.delete(0, 'end')
        container_telefone_motorista.insert(0, novo_texto)
    texto_telefone_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Telefone' )
    texto_telefone_motorista.place(x= 853, y= 20)
    container_telefone_motorista = ctk.CTkEntry(master= dados_motorista, fg_color= 'white', text_color='black')
    container_telefone_motorista.place(x= 850, y= 45)
    container_telefone_motorista.bind("<KeyRelease>", formatar_telefone_motorista)
    

    # inserir Telefone 2 motorista
    def formatar_telefone_2_motorista(event=None):
        texto = container_telefone_2_motorista.get()

        # Remove tudo que não for número
        texto = re.sub(r'\D', '', texto)

        novo_texto = ''
        if len(texto) >= 1:
            novo_texto = '(' + texto[:2]  # DDD
        if len(texto) >= 3:
            if len(texto) >= 11:  # celular com 9 dígitos
                novo_texto += ') ' + texto[2:7] + '-' + texto[7:11]
            elif len(texto) >= 10:  # fixo com 8 dígitos
                novo_texto += ') ' + texto[2:6] + '-' + texto[6:10]
            elif len(texto) > 2:  # ainda incompleto, mas já dá pra exibir parcialmente
                novo_texto += ') ' + texto[2:]

        container_telefone_2_motorista.delete(0, 'end')
        container_telefone_2_motorista.insert(0, novo_texto)
    texto_telefone_2_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Telefone 2' )
    texto_telefone_2_motorista.place(x= 43, y= 90)
    container_telefone_2_motorista = ctk.CTkEntry(master= dados_motorista, fg_color= 'white', text_color='black')
    container_telefone_2_motorista.place(x= 40, y= 115)
    container_telefone_2_motorista.bind("<KeyRelease>", formatar_telefone_2_motorista)
    

    # inserir Telefone 3 motorista
    def formatar_telefone_3_motorista(event=None):
        texto = container_telefone_3_motorista.get()

        # Remove tudo que não for número
        texto = re.sub(r'\D', '', texto)

        novo_texto = ''
        if len(texto) >= 1:
            novo_texto = '(' + texto[:2]  # DDD
        if len(texto) >= 3:
            if len(texto) >= 11:  # celular com 9 dígitos
                novo_texto += ') ' + texto[2:7] + '-' + texto[7:11]
            elif len(texto) >= 10:  # fixo com 8 dígitos
                novo_texto += ') ' + texto[2:6] + '-' + texto[6:10]
            elif len(texto) > 2:  # ainda incompleto, mas já dá pra exibir parcialmente
                novo_texto += ') ' + texto[2:]

        container_telefone_3_motorista.delete(0, 'end')
        container_telefone_3_motorista.insert(0, novo_texto)
    texto_telefone_3_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Telefone 3' )
    texto_telefone_3_motorista.place(x= 193, y= 90)
    container_telefone_3_motorista = ctk.CTkEntry(master= dados_motorista, fg_color= 'white', text_color='black')
    container_telefone_3_motorista.place(x= 190, y= 115)
    container_telefone_3_motorista.bind("<KeyRelease>", formatar_telefone_3_motorista)
    

    # inserir Email do motorista
    texto_email_motorista = ctk.CTkLabel(master= dados_motorista, text= 'Email' )
    texto_email_motorista.place(x= 343, y= 90)
    container_email_motorista = ctk.CTkEntry(master= dados_motorista, fg_color='white', text_color='black')
    container_email_motorista.place(x= 340, y= 115)
    

    # inserir cidade do motorista
    texto_cidade_motorista = ctk.CTkLabel(master=dados_motorista, text='Cidade')
    texto_cidade_motorista.place(x= 493, y= 90)
    container_cidade_motorista = ctk.CTkEntry(master= dados_motorista, fg_color='white', text_color='black')
    container_cidade_motorista.place(x= 490, y= 115)
    

    # inserir bairro do motorista
    texto_bairro_motorista = ctk.CTkLabel(master=dados_motorista, text='Bairro')
    texto_bairro_motorista.place(x= 643, y= 90)
    container_bairro_motorista = ctk.CTkEntry(master= dados_motorista, fg_color='white', text_color='black')
    container_bairro_motorista.place(x= 640, y= 115)
    


    # inserir rua da casa do motorista
    texto_rua_casa_motorista = ctk.CTkLabel(master=dados_motorista, text='Rua')
    texto_rua_casa_motorista.place(x= 793, y= 90)
    container_rua_casa_motorista = ctk.CTkEntry(master= dados_motorista, fg_color='white', text_color='black')
    container_rua_casa_motorista.place(x= 790, y= 115)
    

    # inserir numero da casa do motorista
    def formatar_numero_casa_motorista(event=None):
        texto = container_numero_casa_motorista.get()

        # Remove tudo que não for número
        texto = re.sub(r'\D', '', texto)
        container_numero_casa_motorista.delete(0, 'end')
        container_numero_casa_motorista.insert(0, texto)
    texto_numero_casa_motorista = ctk.CTkLabel(master=dados_motorista, text='Nº Casa')
    texto_numero_casa_motorista.place(x= 43, y= 160)
    container_numero_casa_motorista = ctk.CTkEntry(master= dados_motorista, fg_color='white', text_color='black')
    container_numero_casa_motorista.place(x= 40, y= 185)
    container_numero_casa_motorista.bind('<KeyRelease>', formatar_numero_casa_motorista)
    

    # inserir cep motorista
    def formatar_cep_motorista(event=None):
        texto = container_cep_motorista.get()

        # Remove tudo que não for número
        texto = re.sub(r'\D', '', texto)
        container_cep_motorista.delete(0, 'end')
        container_cep_motorista.insert(0, texto)
    texto_cep_motorista = ctk.CTkLabel(master=dados_motorista, text='CEP')
    texto_cep_motorista.place(x= 193, y= 160)
    container_cep_motorista = ctk.CTkEntry(master= dados_motorista, fg_color='white', text_color='black')
    container_cep_motorista.place(x= 190, y= 185)
    container_cep_motorista.bind('<KeyRelease>', formatar_cep_motorista)
    

    # inserir informações adicionais motorista 
    texto_informacoes_adicionais_motorista = ctk.CTkLabel(master=dados_motorista, text='Informações adicionais')
    texto_informacoes_adicionais_motorista.place(x= 343, y= 160)
    container_informacoes_adicionais_motorista = ctk.CTkTextbox(master= dados_motorista, height= 50, width=140, border_width=2, fg_color='white', text_color='black')
    container_informacoes_adicionais_motorista.place(x= 340, y= 185)
    
    texto_confirmacao = ctk.CTkLabel(master= dados_motorista, text= '')
    texto_confirmacao.place(x= 550, y=215)
    
    # botao adicionar motorista
    def botao_confirmar_app3():
        if len(container_data_cadastro_motorista.get()) < 8:
            texto_confirmacao.configure(text= 'Data de cadastro inválida', text_color = 'red')
        elif container_nome_motorista.get() == '' or None:
            texto_confirmacao.configure(text= 'Insira o nome do motorista', text_color = 'red')
        elif len(container_cpf_motorista.get()) < 11:
            texto_confirmacao.configure(text= 'CPF inválido', text_color = 'red')
        elif len(container_telefone_motorista.get()) < 13:
            texto_confirmacao.configure(text= 'Telefone principal inválido', text_color = 'red')
        else:
            try:
                cursor_mot.execute('''
                INSERT INTO MOTORISTAS (
                    "CÓDIGO", "DATA DO CADASTRO", "NOME", "CPF", "IDENTIDADE",
                    "DATA DO NASCIMENTO", "TELEFONE 1", "TELEFONE 2", "TELEFONE 3", "EMAIL",
                    "CIDADE", "BAIRRO", "RUA", "Nº DA CASA", "CEP", "INFORMAÇÕES ADICIONAIS"
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                container_codigo_motorista.get().upper(),
                container_data_cadastro_motorista.get().upper(),
                container_nome_motorista.get().upper(),
                container_cpf_motorista.get().upper(),
                container_identidade_motorista.get().upper(),
                container_data_nascimento_motorista.get().upper(),
                container_telefone_motorista.get().upper(),
                container_telefone_2_motorista.get().upper(),
                container_telefone_3_motorista.get().upper(),
                container_email_motorista.get().upper(),
                container_cidade_motorista.get().upper(),
                container_bairro_motorista.get().upper(),
                container_rua_casa_motorista.get().upper(),
                container_numero_casa_motorista.get().upper(),
                container_cep_motorista.get().upper(),
                container_informacoes_adicionais_motorista.get("1.0", "end")     
                ))
                banco_mot.commit()
                texto_confirmacao.configure(text= 'Cadastro Confirmado', text_color = 'green')
                
                container_codigo_motorista.delete(0,'end')
                container_data_cadastro_motorista.delete(0,'end')
                container_nome_motorista.delete(0,'end')
                container_cpf_motorista.delete(0,'end')
                container_identidade_motorista.delete(0,'end')
                container_data_nascimento_motorista.delete(0,'end')
                container_telefone_motorista.delete(0,'end')
                container_telefone_2_motorista.delete(0,'end')
                container_telefone_3_motorista.delete(0,'end')
                container_email_motorista.delete(0,'end')
                container_cidade_motorista.delete(0,'end')
                container_bairro_motorista.delete(0,'end')
                container_rua_casa_motorista.delete(0,'end')
                container_numero_casa_motorista.delete(0,'end')
                container_cep_motorista.delete(0,'end')
                container_informacoes_adicionais_motorista.delete('1.0','end')  
                
                container_codigo_motorista.configure(state='normal')
                novo_codigo = gerar_codigo_motorista()
                container_codigo_motorista.delete(0,'end')
                container_codigo_motorista.insert(0, novo_codigo)
                container_codigo_motorista.configure(state='readonly')
                
            except Exception as erro:
                texto_confirmacao.configure(text= f'ERRO: {erro}', text_color = 'red')
            
    botao_confirmar_motorista = ctk.CTkButton(app_3, fg_color='green', image=simbolo_confirmar, text= '', hover_color='dark green', width=80, command=botao_confirmar_app3)
    botao_confirmar_motorista.place(x=800, y=210 )
    
    barra_de_rolagem = tk.Scrollbar()
    
    # cancelar
    def botao_cancelar_app3():
        app_3.destroy()
    botao_cancelar_motorista = ctk.CTkButton(app_3, fg_color='red', image=simbolo_cancelar, text= '', hover_color='dark red', width=80, command=botao_cancelar_app3)
    botao_cancelar_motorista.place(x=900, y=210 )
    
        
botao_cadastrar_motorista = ctk.CTkButton(dados_motorista,image= simbolo_adicionar, text='', height=25,width=80, font=('Arial', 12, 'bold'), fg_color='green', hover_color='dark green', command= cadastrar_motorista)
botao_cadastrar_motorista.place(x=800, y=210 )


################################################################### BOTAO PROCURAR MOTORISTA ####################################################################################################################
def exibir_tabela_mot():
    global app_6
    app_6 = ctk.CTkToplevel(app_2)
    app_6.geometry('1200x400')  # aumentei pra scroll funcionar melhor
    app_6.title('TABELA MOTORISTAS')
    app_6.lift()
    app_6.focus_force()
    app_6.grab_set()

    # Frame principal
    frame_tabela = ctk.CTkFrame(app_6)
    frame_tabela.pack(fill='both', expand=True)

    # Treeview
    tabela_mot = ttk.Treeview(
        frame_tabela,
        columns=(
            "CÓDIGO", "DATA DO CADASTRO", "NOME", "CPF", "IDENTIDADE",
            "DATA DO NASCIMENTO", "TELEFONE 1", "TELEFONE 2", "TELEFONE 3", "EMAIL",
            "CIDADE", "BAIRRO", "RUA", "Nº DA CASA", "CEP", "INFORMAÇÕES ADICIONAIS"
        ),
        show='headings'
    )

    # Configura colunas e cabeçalhos
    for col in tabela_mot["columns"]:
        tabela_mot.heading(col, text=col)
        tabela_mot.column(col, width=150, anchor='center')

    tabela_mot.column("CÓDIGO", width=80)

    # Scrollbars
    scroll_y = ttk.Scrollbar(frame_tabela, orient='vertical', command=tabela_mot.yview)
    scroll_x = ttk.Scrollbar(frame_tabela, orient='horizontal', command=tabela_mot.xview)

    tabela_mot.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

    # Posicionamento com grid
    tabela_mot.grid(row=0, column=0, sticky='nsew')
    scroll_y.grid(row=0, column=1, sticky='ns')
    scroll_x.grid(row=1, column=0, sticky='ew')

    # Expansão automática
    frame_tabela.grid_rowconfigure(0, weight=1)
    frame_tabela.grid_columnconfigure(0, weight=1)

    # Exibe dados se campos nome e cpf estiverem vazios
    if not container_nome_motorista.get() and not container_cpf_motorista.get():
        conexao = sqlite3.connect(caminho_recurso('MOTORISTAS.db'))
        cursor = conexao.cursor()
        cursor.execute('SELECT * FROM MOTORISTAS')
        dados = cursor.fetchall()
        for linha in dados:
            tabela_mot.insert('', 'end', values=linha)
            
    elif len(container_nome_motorista.get()) > 1:
        conexao = sqlite3.connect(caminho_recurso('MOTORISTAS.db'))
        cursor = conexao.cursor()
        container_nome_motorista.get()
        cursor.execute(f'SELECT * FROM MOTORISTAS WHERE NOME LIKE "%{container_nome_motorista.get()}%"')
        dados = cursor.fetchall()
        for linha in dados:
            tabela_mot.insert('', 'end', values=linha)
        
            
            
            
    elif len(container_cpf_motorista.get()) > 1:
        conexao = sqlite3.connect(caminho_recurso('MOTORISTAS.db'))
        cursor = conexao.cursor()
        container_cpf_motorista.get()
        cursor.execute(f'SELECT * FROM MOTORISTAS WHERE NOME LIKE "%{container_cpf_motorista.get()}%"')
        dados = cursor.fetchall()
        for linha in dados:
            tabela_mot.insert('', 'end', values=linha)
        
    
    def adicionar_dados(event):
        selecionado = tabela_mot.focus()
        if selecionado:
            valores = tabela_mot.item(selecionado, 'values')
            if valores:
                container_codigo_motorista.delete(0, 'end')
                container_codigo_motorista.insert(0, valores[0] if len(valores) > 0 else '')
                container_codigo_motorista.configure(state= 'readonly', text_color= 'grey')
                container_data_cadastro_motorista.delete(0, 'end')
                container_data_cadastro_motorista.insert(0, valores[1] if len(valores) > 0 else '')
                container_data_cadastro_motorista.configure(state= 'readonly', text_color= 'grey')
                container_nome_motorista.delete(0, 'end')
                container_nome_motorista.insert(0, valores[2] if len(valores) > 0 else '')
                container_nome_motorista.configure(state= 'readonly', text_color= 'grey')
                container_cpf_motorista.delete(0, 'end')
                container_cpf_motorista.insert(0, valores[3] if len(valores) > 0 else '')
                container_cpf_motorista.configure(state= 'readonly', text_color= 'grey')
                container_identidade_motorista.delete(0, 'end')
                container_identidade_motorista.insert(0, valores[4] if len(valores) > 0 else '')
                container_identidade_motorista.configure(state= 'readonly', text_color= 'grey')
                container_data_nascimento_motorista.delete(0, 'end')
                container_data_nascimento_motorista.insert(0, valores[5] if len(valores) > 0 else '')
                container_data_nascimento_motorista.configure(state= 'readonly', text_color= 'grey')
                container_telefone_motorista.delete(0, 'end')
                container_telefone_motorista.insert(0, valores[6] if len(valores) > 0 else '')
                container_telefone_motorista.configure(state= 'readonly', text_color= 'grey')
                container_telefone_2_motorista.delete(0, 'end')
                container_telefone_2_motorista.insert(0, valores[7] if len(valores) > 0 else '')
                container_telefone_2_motorista.configure(state= 'readonly', text_color= 'grey')
                container_telefone_3_motorista.delete(0, 'end')
                container_telefone_3_motorista.insert(0, valores[8] if len(valores) > 0 else '')
                container_telefone_3_motorista.configure(state= 'readonly', text_color= 'grey')
                container_email_motorista.delete(0, 'end')
                container_email_motorista.insert(0, valores[9] if len(valores) > 0 else '')
                container_email_motorista.configure(state= 'readonly', text_color= 'grey')
                container_cidade_motorista.delete(0, 'end')
                container_cidade_motorista.insert(0, valores[10] if len(valores) > 0 else '')
                container_cidade_motorista.configure(state= 'readonly', text_color= 'grey')
                container_bairro_motorista.delete(0, 'end')
                container_bairro_motorista.insert(0, valores[11] if len(valores) > 0 else '')
                container_bairro_motorista.configure(state= 'readonly', text_color= 'grey')
                container_rua_casa_motorista.delete(0, 'end')
                container_rua_casa_motorista.insert(0, valores[12] if len(valores) > 0 else '')
                container_rua_casa_motorista.configure(state= 'readonly', text_color= 'grey')
                container_numero_casa_motorista.delete(0, 'end')
                container_numero_casa_motorista.insert(0, valores[13] if len(valores) > 0 else '')
                container_numero_casa_motorista.configure(state= 'readonly', text_color= 'grey')
                container_cep_motorista.delete(0, 'end')
                container_cep_motorista.insert(0, valores[14] if len(valores) > 0 else '')
                container_cep_motorista.configure(state= 'readonly', text_color= 'grey')
                container_informacoes_adicionais_motorista.delete('1.0', 'end')
                container_informacoes_adicionais_motorista.insert('1.0', valores[15] if len(valores) > 0 else '')
                container_informacoes_adicionais_motorista.configure(state= 'disabled', text_color= 'grey')

    
    tabela_mot.bind('<Double-1>', adicionar_dados )
    
    
    


botao_procurar_motorista = ctk.CTkButton(dados_motorista, image=simbolo_procurar, text='', height=25, width=80, font=('Arial', 12, 'bold'), command=exibir_tabela_mot)
botao_procurar_motorista.place(x=900, y=210 )



######################################################################################################################################################################################################################


########################## Dados do proprietário ##############################


# Titulo bloco dados do proprietario
titulo_frame_2 = ctk.CTkLabel(app_2, text='Proprietário', font=('Arial', 14, 'bold'))
titulo_frame_2.pack()

# frame de dados do proprietario 
dados_proprietario = ctk.CTkFrame(app_2, width=1050, height=250, border_width=2)
dados_proprietario.pack(pady= 10)

# codigo proprietario
texto_codigo_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Código' )
texto_codigo_proprietario.place(x= 43, y= 20)
container_codigo_proprietario = ctk.CTkEntry(master= dados_proprietario, width=50)
container_codigo_proprietario.place(x= 40, y= 45)

# inserir data do cadastro do proprietario
def formatar_data(event=None):
    texto = container_data_cadastro_proprietario.get()
    texto_limpo = ''.join(filter(str.isdigit, texto))[:8]  # Máximo 8 dígitos

    texto_formatado = ''
    if len(texto_limpo) >= 1:
        texto_formatado += texto_limpo[:2]
    if len(texto_limpo) >= 3:
        texto_formatado = texto_formatado[:2] + '/' + texto_limpo[2:4]
    if len(texto_limpo) >= 5:
        texto_formatado = texto_formatado[:5] + '/' + texto_limpo[4:8]

    container_data_cadastro_proprietario.delete(0, 'end')
    container_data_cadastro_proprietario.insert(0, texto_formatado)
texto_data_cadastro_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Data do Cadastro' )
texto_data_cadastro_proprietario.place(x= 103, y= 20)
container_data_cadastro_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_data_cadastro_proprietario.place(x= 100, y= 45)
container_data_cadastro_proprietario.bind("<KeyRelease>", formatar_data)

# inserir nome do proprietario
texto_nome_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Nome do Proprietário' )
texto_nome_proprietario.place(x= 253, y= 20)
container_nome_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_nome_proprietario.place(x= 250, y= 45)

# inserir CPF proprietario
def formatar_cpf_cnpj(event=None):
    texto = container_cpf_proprietario.get()
    texto = re.sub(r'\D', '', texto)
    texto_limpo = ''.join(filter(str.isdigit, texto))

    if len(texto_limpo) <= 11:  # CPF
        texto_formatado = ''
        if len(texto_limpo) >= 1:
            texto_formatado += texto_limpo[:3]
        if len(texto_limpo) >= 4:
            texto_formatado = texto_formatado[:3] + '.' + texto_limpo[3:6]
        if len(texto_limpo) >= 7:
            texto_formatado = texto_formatado[:7] + '.' + texto_limpo[6:9]
        if len(texto_limpo) >= 10:
            texto_formatado = texto_formatado[:11] + '-' + texto_limpo[9:11]
    else:  # CNPJ
        texto_formatado = ''
        if len(texto_limpo) >= 1:
            texto_formatado += texto_limpo[:2]
        if len(texto_limpo) >= 3:
            texto_formatado = texto_formatado[:2] + '.' + texto_limpo[2:5]
        if len(texto_limpo) >= 6:
            texto_formatado = texto_formatado[:6] + '.' + texto_limpo[5:8]
        if len(texto_limpo) >= 9:
            texto_formatado = texto_formatado[:10] + '/' + texto_limpo[8:12]
        if len(texto_limpo) >= 13:
            texto_formatado = texto_formatado[:15] + '-' + texto_limpo[12:14]

    container_cpf_proprietario.delete(0, 'end')
    container_cpf_proprietario.insert(0, texto_formatado)
    
texto_cpf_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'CPF/CNPJ' )
texto_cpf_proprietario.place(x= 403, y= 20)
container_cpf_proprietario = ctk.CTkEntry(master=dados_proprietario)
container_cpf_proprietario.place(x=400, y=45)
container_cpf_proprietario.bind("<KeyRelease>", formatar_cpf_cnpj)


# inserir Identidade proprietario
def formatar_identidade_proprietario(event=None):
    texto = container_identidade_proprietario.get()

    # Remove tudo que não for número
    texto = re.sub(r'\D', '', texto)
    container_identidade_proprietario.delete(0, 'end')
    container_identidade_proprietario.insert(0, texto)
    
texto_identidade_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Identidade' )
texto_identidade_proprietario.place(x= 553, y= 20)
container_identidade_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_identidade_proprietario.place(x= 550, y= 45)
container_identidade_proprietario.bind("<KeyRelease>", formatar_identidade_proprietario)


# inserir Data Nascimento proprietario
def formatar_data_nascimento(event=None):
    texto = container_data_nascimento_proprietario.get()
    texto_limpo = ''.join(filter(str.isdigit, texto))[:8]  # Máximo 8 dígitos

    texto_formatado = ''
    if len(texto_limpo) >= 1:
        texto_formatado += texto_limpo[:2]
    if len(texto_limpo) >= 3:
        texto_formatado = texto_formatado[:2] + '/' + texto_limpo[2:4]
    if len(texto_limpo) >= 5:
        texto_formatado = texto_formatado[:5] + '/' + texto_limpo[4:8]

    container_data_nascimento_proprietario.delete(0, 'end')
    container_data_nascimento_proprietario.insert(0, texto_formatado)
texto_data_nascimento_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Data Nascimento' )
texto_data_nascimento_proprietario.place(x= 703, y= 20)
container_data_nascimento_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_data_nascimento_proprietario.place(x= 700, y= 45)
container_data_nascimento_proprietario.bind("<KeyRelease>", formatar_data_nascimento)

def formatar_telefone_proprietario(event=None):
    texto = container_telefone_proprietario.get()

    # Remove tudo que não for número
    texto = re.sub(r'\D', '', texto)

    novo_texto = ''
    if len(texto) >= 1:
        novo_texto = '(' + texto[:2]  # DDD
    if len(texto) >= 3:
        if len(texto) >= 11:  # celular com 9 dígitos
            novo_texto += ') ' + texto[2:7] + '-' + texto[7:11]
        elif len(texto) >= 10:  # fixo com 8 dígitos
            novo_texto += ') ' + texto[2:6] + '-' + texto[6:10]
        elif len(texto) > 2:  # ainda incompleto, mas já dá pra exibir parcialmente
            novo_texto += ') ' + texto[2:]

    container_telefone_proprietario.delete(0, 'end')
    container_telefone_proprietario.insert(0, novo_texto)
texto_telefone_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Telefone' )
texto_telefone_proprietario.place(x= 853, y= 20)
container_telefone_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_telefone_proprietario.place(x= 850, y= 45)
container_telefone_proprietario.bind("<KeyRelease>", formatar_telefone_proprietario)

# inserir Telefone 2 proprietario
def formatar_telefone_2_proprietario(event=None):
    texto = container_telefone_2_proprietario.get()

    # Remove tudo que não for número
    texto = re.sub(r'\D', '', texto)

    novo_texto = ''
    if len(texto) >= 1:
        novo_texto = '(' + texto[:2]  # DDD
    if len(texto) >= 3:
        if len(texto) >= 11:  # celular com 9 dígitos
            novo_texto += ') ' + texto[2:7] + '-' + texto[7:11]
        elif len(texto) >= 10:  # fixo com 8 dígitos
            novo_texto += ') ' + texto[2:6] + '-' + texto[6:10]
        elif len(texto) > 2:  # ainda incompleto, mas já dá pra exibir parcialmente
            novo_texto += ') ' + texto[2:]

    container_telefone_2_proprietario.delete(0, 'end')
    container_telefone_2_proprietario.insert(0, novo_texto)
texto_telefone_2_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Telefone 2' )
texto_telefone_2_proprietario.place(x= 43, y= 90)
container_telefone_2_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_telefone_2_proprietario.place(x= 40, y= 115)
container_telefone_2_proprietario.bind("<KeyRelease>", formatar_telefone_2_proprietario)

# inserir Telefone 3 proprietario
def formatar_telefone_3_proprietario(event=None):
    texto = container_telefone_3_proprietario.get()

    # Remove tudo que não for número
    texto = re.sub(r'\D', '', texto)

    novo_texto = ''
    if len(texto) >= 1:
        novo_texto = '(' + texto[:2]  # DDD
    if len(texto) >= 3:
        if len(texto) >= 11:  # celular com 9 dígitos
            novo_texto += ') ' + texto[2:7] + '-' + texto[7:11]
        elif len(texto) >= 10:  # fixo com 8 dígitos
            novo_texto += ') ' + texto[2:6] + '-' + texto[6:10]
        elif len(texto) > 2:  # ainda incompleto, mas já dá pra exibir parcialmente
            novo_texto += ') ' + texto[2:]

    container_telefone_3_proprietario.delete(0, 'end')
    container_telefone_3_proprietario.insert(0, novo_texto)
texto_telefone_3_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Telefone 3' )
texto_telefone_3_proprietario.place(x= 193, y= 90)
container_telefone_3_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_telefone_3_proprietario.place(x= 190, y= 115)
container_telefone_3_proprietario.bind("<KeyRelease>", formatar_telefone_3_proprietario)

# inserir Email do proprietario
texto_email_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Email' )
texto_email_proprietario.place(x= 343, y= 90)
container_email_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_email_proprietario.place(x= 340, y= 115)

# inserir cidade do proprietario
texto_cidade_proprietario = ctk.CTkLabel(master=dados_proprietario, text='Cidade')
texto_cidade_proprietario.place(x= 493, y= 90)
container_cidade_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_cidade_proprietario.place(x= 490, y= 115)

# inserir bairro do proprietario
texto_bairro_proprietario = ctk.CTkLabel(master=dados_proprietario, text='Bairro')
texto_bairro_proprietario.place(x= 643, y= 90)
container_bairro_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_bairro_proprietario.place(x= 640, y= 115)


# inserir rua da casa do proprietario
texto_rua_casa_proprietario = ctk.CTkLabel(master=dados_proprietario, text='Rua')
texto_rua_casa_proprietario.place(x= 793, y= 90)
container_rua_casa_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_rua_casa_proprietario.place(x= 790, y= 115)

# inserir numero da casa do proprietario
def formatar_numero_casa_proprietario(event=None):
    texto = container_numero_casa_proprietario.get()

    # Remove tudo que não for número
    texto = re.sub(r'\D', '', texto)
    container_numero_casa_proprietario.delete(0, 'end')
    container_numero_casa_proprietario.insert(0, texto)
    
texto_numero_casa_proprietario = ctk.CTkLabel(master=dados_proprietario, text='Nº Casa')
texto_numero_casa_proprietario.place(x= 43, y= 160)
container_numero_casa_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_numero_casa_proprietario.place(x= 40, y= 185)
container_numero_casa_proprietario.bind("<KeyRelease>", formatar_numero_casa_proprietario)

# inserir cep proprietario
def formatar_cep_proprietario(event=None):
    texto = container_cep_proprietario.get()

    # Remove tudo que não for número
    texto = re.sub(r'\D', '', texto)
    container_cep_proprietario.delete(0, 'end')
    container_cep_proprietario.insert(0, texto)
texto_cep_proprietario = ctk.CTkLabel(master=dados_proprietario, text='CEP')
texto_cep_proprietario.place(x= 193, y= 160)
container_cep_proprietario = ctk.CTkEntry(master= dados_proprietario)
container_cep_proprietario.place(x= 190, y= 185)
container_cep_proprietario.bind("<KeyRelease>", formatar_cep_proprietario)

# inserir informações adicionais proprietario 
texto_informacoes_adicionais_proprietario = ctk.CTkLabel(master=dados_proprietario, text='Informações adicionais')
texto_informacoes_adicionais_proprietario.place(x= 343, y= 160)
container_informacoes_adicionais_proprietario = ctk.CTkTextbox(master= dados_proprietario, height= 50, width=140, border_width=2)
container_informacoes_adicionais_proprietario.place(x= 340, y= 185)


def limpar_conteudo():
    container_codigo_proprietario.configure(state='normal')
    container_codigo_proprietario.delete(0, 'end')
    container_data_cadastro_proprietario.configure(state='normal')
    container_data_cadastro_proprietario.delete(0, 'end')
    container_nome_proprietario.configure(state='normal')
    container_nome_proprietario.delete(0, 'end')
    container_cpf_proprietario.configure(state='normal')
    container_cpf_proprietario.delete(0, 'end')
    container_identidade_proprietario.configure(state='normal')
    container_identidade_proprietario.delete(0, 'end')
    container_data_nascimento_proprietario.configure(state='normal')
    container_data_nascimento_proprietario.delete(0, 'end')
    container_telefone_proprietario.configure(state='normal')
    container_telefone_proprietario.delete(0, 'end')
    container_telefone_2_proprietario.configure(state='normal')
    container_telefone_2_proprietario.delete(0, 'end')
    container_telefone_3_proprietario.configure(state='normal')
    container_telefone_3_proprietario.delete(0, 'end')
    container_email_proprietario.configure(state='normal')
    container_email_proprietario.delete(0, 'end')
    container_cidade_proprietario.configure(state='normal')
    container_cidade_proprietario.delete(0, 'end')
    container_bairro_proprietario.configure(state='normal')
    container_bairro_proprietario.delete(0, 'end')
    container_rua_casa_proprietario.configure(state='normal')
    container_rua_casa_proprietario.delete(0, 'end')
    container_numero_casa_proprietario.configure(state='normal')
    container_numero_casa_proprietario.delete(0, 'end')
    container_cep_proprietario.configure(state='normal')
    container_cep_proprietario.delete(0, 'end')
    container_informacoes_adicionais_proprietario.configure(state='normal')
    container_informacoes_adicionais_proprietario.delete("1.0", "end")     
    
botao_limpar = ctk.CTkButton(dados_proprietario, text='Limpar', fg_color='light yellow', font=('Arial', 10), text_color='black', width=30, hover_color='grey', command=limpar_conteudo)
botao_limpar.place(x=1000, y=4)

############################################################ BOTÃO CADASTRAR PROPRIETARIOS #####################################################################################################################

def cadastrar_proprietario():
    app_4 = ctk.CTkToplevel(app_2)
    app_4.lift()
    app_4.focus_force()
    app_4.grab_set()
    app_4.maxsize(width=1000, height=300)
    app_4.title('Cadastrar Proprietario')
    codigo_proprietario = gerar_codigo_proprietario()
    # frame de dados do proprietario 
    dados_proprietario = ctk.CTkFrame(app_4, width=1050, height=250, border_width=2)
    dados_proprietario.pack(pady= 10)
    # codigo do proprietario
    texto_codigo_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Código' )
    texto_codigo_proprietario.place(x= 43, y= 20)
    container_codigo_proprietario = ctk.CTkEntry(master= dados_proprietario, width=50)
    container_codigo_proprietario.insert(0, codigo_proprietario)
    container_codigo_proprietario.configure(state='readonly', text_color= 'grey')
    container_codigo_proprietario.place(x= 40, y= 45)
    

    # inserir data do cadastro do proprietario
    def formatar_data(event=None):
        texto = container_data_cadastro_proprietario.get()
        texto_limpo = ''.join(filter(str.isdigit, texto))[:8]  # Máximo 8 dígitos

        texto_formatado = ''
        if len(texto_limpo) >= 1:
            texto_formatado += texto_limpo[:2]
        if len(texto_limpo) >= 3:
            texto_formatado = texto_formatado[:2] + '/' + texto_limpo[2:4]
        if len(texto_limpo) >= 5:
            texto_formatado = texto_formatado[:5] + '/' + texto_limpo[4:8]

        container_data_cadastro_proprietario.delete(0, 'end')
        container_data_cadastro_proprietario.insert(0, texto_formatado)
    texto_data_cadastro_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Data do Cadastro' )
    texto_data_cadastro_proprietario.place(x= 103, y= 20)
    container_data_cadastro_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color='white', text_color='black')
    container_data_cadastro_proprietario.place(x= 100, y= 45)
    container_data_cadastro_proprietario.bind("<KeyRelease>", formatar_data)
    

    # inserir nome do proprietario
    texto_nome_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Nome do proprietario')
    texto_nome_proprietario.place(x= 253, y= 20)
    container_nome_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color='white', text_color='black')
    container_nome_proprietario.place(x= 250, y= 45)
    

    # inserir CPF proprietario
    def formatar_cpf_cnpj(event=None):
        texto = container_cpf_proprietario.get()
        texto_limpo = ''.join(filter(str.isdigit, texto))

        texto = re.sub(r'\D', '', texto)

        if len(texto_limpo) <= 11:  # CPF
            texto_formatado = ''
            if len(texto_limpo) >= 1:
                texto_formatado += texto_limpo[:3]
            if len(texto_limpo) >= 4:
                texto_formatado = texto_formatado[:3] + '.' + texto_limpo[3:6]
            if len(texto_limpo) >= 7:
                texto_formatado = texto_formatado[:7] + '.' + texto_limpo[6:9]
            if len(texto_limpo) >= 10:
                texto_formatado = texto_formatado[:11] + '-' + texto_limpo[9:11]
        else:  # CNPJ
            texto_formatado = ''
            if len(texto_limpo) >= 1:
                texto_formatado += texto_limpo[:2]
            if len(texto_limpo) >= 3:
                texto_formatado = texto_formatado[:2] + '.' + texto_limpo[2:5]
            if len(texto_limpo) >= 6:
                texto_formatado = texto_formatado[:6] + '.' + texto_limpo[5:8]
            if len(texto_limpo) >= 9:
                texto_formatado = texto_formatado[:10] + '/' + texto_limpo[8:12]
            if len(texto_limpo) >= 13:
                texto_formatado = texto_formatado[:15] + '-' + texto_limpo[12:14]

        container_cpf_proprietario.delete(0, 'end')
        container_cpf_proprietario.insert(0, texto_formatado)
        
    texto_cpf_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'CPF/CNPJ' )
    texto_cpf_proprietario.place(x= 403, y= 20)
    container_cpf_proprietario = ctk.CTkEntry(master=dados_proprietario, text_color='black', fg_color='white')
    container_cpf_proprietario.place(x=400, y=45)
    container_cpf_proprietario.bind("<KeyRelease>", formatar_cpf_cnpj)
    


    # inserir Identidade do proprietario
    def formatar_identidade_proprietario(event=None):
        texto = container_identidade_proprietario.get()

        # Remove tudo que não for número
        texto = re.sub(r'\D', '', texto)
        container_identidade_proprietario.delete(0, 'end')
        container_identidade_proprietario.insert(0, texto)
        
    texto_identidade_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Identidade' )
    texto_identidade_proprietario.place(x= 553, y= 20)
    container_identidade_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color='white', text_color='black')
    container_identidade_proprietario.place(x= 550, y= 45)
    container_cpf_proprietario.bind("<KeyRelease>", formatar_identidade_proprietario)
    


    # inserir Data Nascimento do proprietario
    def formatar_data_nascimento(event=None):
        texto = container_data_nascimento_proprietario.get()
        texto_limpo = ''.join(filter(str.isdigit, texto))[:8]  # Máximo 8 dígitos

        texto_formatado = ''
        if len(texto_limpo) >= 1:
            texto_formatado += texto_limpo[:2]
        if len(texto_limpo) >= 3:
            texto_formatado = texto_formatado[:2] + '/' + texto_limpo[2:4]
        if len(texto_limpo) >= 5:
            texto_formatado = texto_formatado[:5] + '/' + texto_limpo[4:8]

        container_data_nascimento_proprietario.delete(0, 'end')
        container_data_nascimento_proprietario.insert(0, texto_formatado)
    texto_data_nascimento_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Data Nascimento' )
    texto_data_nascimento_proprietario.place(x= 703, y= 20)
    container_data_nascimento_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color='white', text_color='black')
    container_data_nascimento_proprietario.place(x= 700, y= 45)
    container_data_nascimento_proprietario.bind("<KeyRelease>", formatar_data_nascimento)
    

    def formatar_telefone_proprietario(event=None):
        texto = container_telefone_proprietario.get()

        # Remove tudo que não for número
        texto = re.sub(r'\D', '', texto)

        novo_texto = ''
        if len(texto) >= 1:
            novo_texto = '(' + texto[:2]  # DDD
        if len(texto) >= 3:
            if len(texto) >= 11:  # celular com 9 dígitos
                novo_texto += ') ' + texto[2:7] + '-' + texto[7:11]
            elif len(texto) >= 10:  # fixo com 8 dígitos
                novo_texto += ') ' + texto[2:6] + '-' + texto[6:10]
            elif len(texto) > 2:  # ainda incompleto, mas já dá pra exibir parcialmente
                novo_texto += ') ' + texto[2:]

        container_telefone_proprietario.delete(0, 'end')
        container_telefone_proprietario.insert(0, novo_texto)
    texto_telefone_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Telefone' )
    texto_telefone_proprietario.place(x= 853, y= 20)
    container_telefone_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color= 'white', text_color='black')
    container_telefone_proprietario.place(x= 850, y= 45)
    container_telefone_proprietario.bind("<KeyRelease>", formatar_telefone_proprietario)
    

    # inserir Telefone 2 proprietario
    def formatar_telefone_2_proprietario(event=None):
        texto = container_telefone_2_proprietario.get()

        # Remove tudo que não for número
        texto = re.sub(r'\D', '', texto)

        novo_texto = ''
        if len(texto) >= 1:
            novo_texto = '(' + texto[:2]  # DDD
        if len(texto) >= 3:
            if len(texto) >= 11:  # celular com 9 dígitos
                novo_texto += ') ' + texto[2:7] + '-' + texto[7:11]
            elif len(texto) >= 10:  # fixo com 8 dígitos
                novo_texto += ') ' + texto[2:6] + '-' + texto[6:10]
            elif len(texto) > 2:  # ainda incompleto, mas já dá pra exibir parcialmente
                novo_texto += ') ' + texto[2:]

        container_telefone_2_proprietario.delete(0, 'end')
        container_telefone_2_proprietario.insert(0, novo_texto)
    texto_telefone_2_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Telefone 2' )
    texto_telefone_2_proprietario.place(x= 43, y= 90)
    container_telefone_2_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color= 'white', text_color='black')
    container_telefone_2_proprietario.place(x= 40, y= 115)
    container_telefone_2_proprietario.bind("<KeyRelease>", formatar_telefone_2_proprietario)
    

    # inserir Telefone 3 proprietario
    def formatar_telefone_3_proprietario(event=None):
        texto = container_telefone_3_proprietario.get()

        # Remove tudo que não for número
        texto = re.sub(r'\D', '', texto)

        novo_texto = ''
        if len(texto) >= 1:
            novo_texto = '(' + texto[:2]  # DDD
        if len(texto) >= 3:
            if len(texto) >= 11:  # celular com 9 dígitos
                novo_texto += ') ' + texto[2:7] + '-' + texto[7:11]
            elif len(texto) >= 10:  # fixo com 8 dígitos
                novo_texto += ') ' + texto[2:6] + '-' + texto[6:10]
            elif len(texto) > 2:  # ainda incompleto, mas já dá pra exibir parcialmente
                novo_texto += ') ' + texto[2:]

        container_telefone_3_proprietario.delete(0, 'end')
        container_telefone_3_proprietario.insert(0, novo_texto)
    texto_telefone_3_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Telefone 3' )
    texto_telefone_3_proprietario.place(x= 193, y= 90)
    container_telefone_3_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color= 'white', text_color='black')
    container_telefone_3_proprietario.place(x= 190, y= 115)
    container_telefone_3_proprietario.bind("<KeyRelease>", formatar_telefone_3_proprietario)
    

    # inserir Email do proprietario
    texto_email_proprietario = ctk.CTkLabel(master= dados_proprietario, text= 'Email' )
    texto_email_proprietario.place(x= 343, y= 90)
    container_email_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color='white', text_color='black')
    container_email_proprietario.place(x= 340, y= 115)
    

    # inserir cidade do proprietario
    texto_cidade_proprietario = ctk.CTkLabel(master=dados_proprietario, text='Cidade')
    texto_cidade_proprietario.place(x= 493, y= 90)
    container_cidade_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color='white', text_color='black')
    container_cidade_proprietario.place(x= 490, y= 115)
    

    # inserir bairro do proprietario
    texto_bairro_proprietario = ctk.CTkLabel(master=dados_proprietario, text='Bairro')
    texto_bairro_proprietario.place(x= 643, y= 90)
    container_bairro_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color='white', text_color='black')
    container_bairro_proprietario.place(x= 640, y= 115)
    


    # inserir rua da casa do proprietario
    texto_rua_casa_proprietario = ctk.CTkLabel(master=dados_proprietario, text='Rua')
    texto_rua_casa_proprietario.place(x= 793, y= 90)
    container_rua_casa_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color='white', text_color='black')
    container_rua_casa_proprietario.place(x= 790, y= 115)
    

    # inserir numero da casa do proprietario
    def formatar_numero_casa_proprietario(event=None):
        texto = container_numero_casa_proprietario.get()

        # Remove tudo que não for número
        texto = re.sub(r'\D', '', texto)
        container_numero_casa_proprietario.delete(0, 'end')
        container_numero_casa_proprietario.insert(0, texto)
        
    texto_numero_casa_proprietario = ctk.CTkLabel(master=dados_proprietario, text='Nº Casa')
    texto_numero_casa_proprietario.place(x= 43, y= 160)
    container_numero_casa_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color='white', text_color='black')
    container_numero_casa_proprietario.place(x= 40, y= 185)
    container_numero_casa_proprietario.bind("<KeyRelease>", formatar_numero_casa_proprietario)
    

    # inserir cep proprietario
    def formatar_cep_proprietario(event=None):
        texto = container_cep_proprietario.get()

        # Remove tudo que não for número
        texto = re.sub(r'\D', '', texto)
        container_cep_proprietario.delete(0, 'end')
        container_cep_proprietario.insert(0, texto)
        
    texto_cep_proprietario = ctk.CTkLabel(master=dados_proprietario, text='CEP')
    texto_cep_proprietario.place(x= 193, y= 160)
    container_cep_proprietario = ctk.CTkEntry(master= dados_proprietario, fg_color='white', text_color='black')
    container_cep_proprietario.place(x= 190, y= 185)
    container_cep_proprietario.bind("<KeyRelease>", formatar_cep_proprietario)
    

    # inserir informações adicionais proprietario 
    texto_informacoes_adicionais_proprietario = ctk.CTkLabel(master=dados_proprietario, text='Informações adicionais')
    texto_informacoes_adicionais_proprietario.place(x= 343, y= 160)
    container_informacoes_adicionais_proprietario = ctk.CTkTextbox(master= dados_proprietario, height= 50, width=140, border_width=2, fg_color='white', text_color='black')
    container_informacoes_adicionais_proprietario.place(x= 340, y= 185)
    
    texto_confirmacao = ctk.CTkLabel(master= dados_proprietario, text= '')
    texto_confirmacao.place(x= 550, y=215)
    
    # botao adicionar proprietario
    def botao_confirmar_app4():
        if len(container_data_cadastro_proprietario.get()) < 8:
            texto_confirmacao.configure(text= 'Data de cadastro inválida', text_color = 'red')
        elif container_nome_proprietario.get() == '' or None:
            texto_confirmacao.configure(text= 'Insira o nome do proprietário', text_color = 'red')
        elif len(container_cpf_proprietario.get()) < 11:
            texto_confirmacao.configure(text= 'CPF inválido', text_color = 'red')
        elif len(container_telefone_proprietario.get()) < 13:
            texto_confirmacao.configure(text= 'Telefone principal inválido', text_color = 'red')
        else:
            try:
                cursor_prop.execute('''
                INSERT INTO PROPRIETARIOS (
                    "CÓDIGO", "DATA DO CADASTRO", "NOME", "CPF/CNPJ", "IDENTIDADE",
                    "DATA DO NASCIMENTO", "TELEFONE 1", "TELEFONE 2", "TELEFONE 3", "EMAIL",
                    "CIDADE", "BAIRRO", "RUA", "Nº DA CASA", "CEP", "INFORMAÇÕES ADICIONAIS"
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                container_codigo_proprietario.get().upper(),
                container_data_cadastro_proprietario.get().upper(),
                container_nome_proprietario.get().upper(),
                container_cpf_proprietario.get().upper(),
                container_identidade_proprietario.get().upper(),
                container_data_nascimento_proprietario.get().upper(),
                container_telefone_proprietario.get().upper(),
                container_telefone_2_proprietario.get().upper(),
                container_telefone_3_proprietario.get().upper(),
                container_email_proprietario.get().upper(),
                container_cidade_proprietario.get().upper(),
                container_bairro_proprietario.get().upper(),
                container_rua_casa_proprietario.get().upper(),
                container_numero_casa_proprietario.get().upper(),
                container_cep_proprietario.get().upper(),
                container_informacoes_adicionais_proprietario.get("1.0", "end")     
                ))
                banco_prop.commit()
                texto_confirmacao.configure(text= 'Cadastro Confirmado', text_color = 'green')
                
                container_codigo_proprietario.delete(0,'end')
                container_data_cadastro_proprietario.delete(0,'end')
                container_nome_proprietario.delete(0,'end')
                container_cpf_proprietario.delete(0,'end')
                container_identidade_proprietario.delete(0,'end')
                container_data_nascimento_proprietario.delete(0,'end')
                container_telefone_proprietario.delete(0,'end')
                container_telefone_2_proprietario.delete(0,'end')
                container_telefone_3_proprietario.delete(0,'end')
                container_email_proprietario.delete(0,'end')
                container_cidade_proprietario.delete(0,'end')
                container_bairro_proprietario.delete(0,'end')
                container_rua_casa_proprietario.delete(0,'end')
                container_numero_casa_proprietario.delete(0,'end')
                container_cep_proprietario.delete(0,'end')
                container_informacoes_adicionais_proprietario.delete('1.0','end')  
                
                container_codigo_proprietario.configure(state='normal')
                novo_codigo = gerar_codigo_proprietario()
                container_codigo_proprietario.delete(0,'end')
                container_codigo_proprietario.insert(0, novo_codigo)
                container_codigo_proprietario.configure(state='readonly')
                
            except Exception as erro:
                texto_confirmacao.configure(text= f'ERRO: {erro}', text_color = 'red')
        
    botao_confirmar_proprietario = ctk.CTkButton(app_4, fg_color='green', image=simbolo_confirmar, text= '', hover_color='dark green', width=80, command= botao_confirmar_app4)
    botao_confirmar_proprietario.place(x=800, y=210 )
    # cancelar
    def botao_cancelar_app4():
        app_4.destroy()
    botao_cancelar_proprietario = ctk.CTkButton(app_4, fg_color='red', image=simbolo_cancelar, text= '', hover_color='dark red', width=80, command=botao_cancelar_app4)
    botao_cancelar_proprietario.place(x=900, y=210 )
    
        
botao_cadastrar_proprietario = ctk.CTkButton(dados_proprietario,image= simbolo_adicionar, text='', height=25,width=80, font=('Arial', 12, 'bold'), fg_color='green', hover_color='dark green', command= cadastrar_proprietario)
botao_cadastrar_proprietario.place(x=800, y=210 )

# ################################################################### BOTAO PROCURAR PROPRIETARIO ####################################################################################################################

def exibir_tabela_prop():
    global app_7
    app_7 = ctk.CTkToplevel(app_2)
    app_7.geometry('1200x400')
    app_7.title('TABELA PROPRIETÁRIOS')
    app_7.lift()
    app_7.focus_force()
    app_7.grab_set()

    # Frame principal
    frame_tabela = ctk.CTkFrame(app_7)
    frame_tabela.pack(fill='both', expand=True)

    # Treeview
    tabela_prop = ttk.Treeview(
        frame_tabela,
        columns=(
            "CÓDIGO", "DATA DO CADASTRO", "NOME", "CPF/CNPJ", "IDENTIDADE",
            "DATA DO NASCIMENTO", "TELEFONE 1", "TELEFONE 2", "TELEFONE 3", "EMAIL",
            "CIDADE", "BAIRRO", "RUA", "Nº DA CASA", "CEP", "INFORMAÇÕES ADICIONAIS"
        ),
        show='headings'
    )

    # Configura colunas e cabeçalhos
    for col in tabela_prop["columns"]:
        tabela_prop.heading(col, text=col)
        tabela_prop.column(col, width=150, anchor='center')

    tabela_prop.column("CÓDIGO", width=80)

    # Scrollbars
    scroll_y = ttk.Scrollbar(frame_tabela, orient='vertical', command=tabela_prop.yview)
    scroll_x = ttk.Scrollbar(frame_tabela, orient='horizontal', command=tabela_prop.xview)

    tabela_prop.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

    # Posicionamento com grid
    tabela_prop.grid(row=0, column=0, sticky='nsew')
    scroll_y.grid(row=0, column=1, sticky='ns')
    scroll_x.grid(row=1, column=0, sticky='ew')

    # Expansão automática
    frame_tabela.grid_rowconfigure(0, weight=1)
    frame_tabela.grid_columnconfigure(0, weight=1)

    # Exibe dados se campos nome e cpf estiverem vazios

    if not container_nome_proprietario.get() and not container_cpf_proprietario.get():
        conexao = sqlite3.connect(caminho_recurso('PROPRIETARIOS.db'))
        cursor = conexao.cursor()
        cursor.execute('SELECT * FROM PROPRIETARIOS')
        dados = cursor.fetchall()
        for linha in dados:
            tabela_prop.insert('', 'end', values=linha)
            
    elif len(container_nome_proprietario.get()) > 1:
        conexao = sqlite3.connect(caminho_recurso('PROPRIETARIOS.db'))
        cursor = conexao.cursor()
        container_nome_proprietario.get()
        cursor.execute(f'SELECT * FROM PROPRIETARIOS WHERE NOME LIKE "%{container_nome_proprietario.get()}%"')
        dados = cursor.fetchall()
        for linha in dados:
            tabela_prop.insert('', 'end', values=linha)
        
            
            
            
    elif len(container_cpf_proprietario.get()) > 1:
        conexao = sqlite3.connect(caminho_recurso('PROPRIETARIOS.db'))
        cursor = conexao.cursor()
        container_cpf_proprietario.get()
        cursor.execute(f'SELECT * FROM PROPRIETARIOS WHERE NOME LIKE "%{container_cpf_proprietario.get()}%"')
        dados = cursor.fetchall()
        for linha in dados:
            tabela_prop.insert('', 'end', values=linha)
    
    def adicionar_dados(event):
        selecionado = tabela_prop.focus()
        if selecionado:
            valores = tabela_prop.item(selecionado, 'values')
            if valores:
                container_codigo_proprietario.delete(0, 'end')
                container_codigo_proprietario.insert(0, valores[0] if len(valores) > 0 else '')
                container_codigo_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_data_cadastro_proprietario.delete(0, 'end')
                container_data_cadastro_proprietario.insert(0, valores[1] if len(valores) > 0 else '')
                container_data_cadastro_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_nome_proprietario.delete(0, 'end')
                container_nome_proprietario.insert(0, valores[2] if len(valores) > 0 else '')
                container_nome_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_cpf_proprietario.delete(0, 'end')
                container_cpf_proprietario.insert(0, valores[3] if len(valores) > 0 else '')
                container_cpf_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_identidade_proprietario.delete(0, 'end')
                container_identidade_proprietario.insert(0, valores[4] if len(valores) > 0 else '')
                container_identidade_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_data_nascimento_proprietario.delete(0, 'end')
                container_data_nascimento_proprietario.insert(0, valores[5] if len(valores) > 0 else '')
                container_data_nascimento_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_telefone_proprietario.delete(0, 'end')
                container_telefone_proprietario.insert(0, valores[6] if len(valores) > 0 else '')
                container_telefone_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_telefone_2_proprietario.delete(0, 'end')
                container_telefone_2_proprietario.insert(0, valores[7] if len(valores) > 0 else '')
                container_telefone_2_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_telefone_3_proprietario.delete(0, 'end')
                container_telefone_3_proprietario.insert(0, valores[8] if len(valores) > 0 else '')
                container_telefone_3_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_email_proprietario.delete(0, 'end')
                container_email_proprietario.insert(0, valores[9] if len(valores) > 0 else '')
                container_email_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_cidade_proprietario.delete(0, 'end')
                container_cidade_proprietario.insert(0, valores[10] if len(valores) > 0 else '')
                container_cidade_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_bairro_proprietario.delete(0, 'end')
                container_bairro_proprietario.insert(0, valores[11] if len(valores) > 0 else '')
                container_bairro_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_rua_casa_proprietario.delete(0, 'end')
                container_rua_casa_proprietario.insert(0, valores[12] if len(valores) > 0 else '')
                container_rua_casa_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_numero_casa_proprietario.delete(0, 'end')
                container_numero_casa_proprietario.insert(0, valores[13] if len(valores) > 0 else '')
                container_numero_casa_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_cep_proprietario.delete(0, 'end')
                container_cep_proprietario.insert(0, valores[14] if len(valores) > 0 else '')
                container_cep_proprietario.configure(state= 'readonly', text_color= 'grey')
                container_informacoes_adicionais_proprietario.delete('1.0', 'end')
                container_informacoes_adicionais_proprietario.insert('1.0', valores[15] if len(valores) > 0 else '')
                container_informacoes_adicionais_proprietario.configure(state= 'disabled', text_color= 'grey')

        
    tabela_prop.bind('<Double-1>', adicionar_dados)
    

botao_procurar_proprietario = ctk.CTkButton(dados_proprietario, image=simbolo_procurar, text='', height=25, width=80, font=('Arial', 12, 'bold'), command=exibir_tabela_prop)
botao_procurar_proprietario.place(x=900, y=210 )
#########################################################################################################################################################################################################################


########################## frame lado a lado ##############################


# container para separar lado a lado veiculo e extras

frame_titulos = ctk.CTkFrame(app_2, height=30, width=1500, fg_color='transparent')
frame_titulos.pack()

frame_veiculo_extras = ctk.CTkFrame(app_2, fg_color='transparent')
frame_veiculo_extras.pack(pady=10)

########################## Dados do veiculo ##############################

# Titulo bloco dados veiculo
titulo_frame_3 =  ctk.CTkLabel(frame_titulos, font=('Arial', 14, 'bold'), text='Dados Veículo')
titulo_frame_3.place(x=200, y=0)

# Frame dados veiculo
dados_veiculo = ctk.CTkFrame(frame_veiculo_extras, width=580, height=200, border_width=2 )
dados_veiculo.pack(padx=10, side= 'left')


# codigo veiculo
texto_codigo_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Código' )
texto_codigo_veiculo.place(x= 43, y= 20)
container_codigo_veiculo = ctk.CTkEntry(master= dados_veiculo, width=50)
container_codigo_veiculo.place(x= 40, y= 45)

# inserir marca do veiculo
texto_marca_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Marca' )
texto_marca_veiculo.place(x= 103, y= 20)
container_marca_veiculo = ctk.CTkEntry(master= dados_veiculo)
container_marca_veiculo.place(x= 100, y= 45)

# inserir modelo do veiculo
texto_modelo_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Modelo' )
texto_modelo_veiculo.place(x= 253, y= 20)
container_modelo_veiculo = ctk.CTkEntry(master= dados_veiculo)
container_modelo_veiculo.place(x= 250, y= 45)

# inserir ano veiculo
texto_ano_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Ano' )
texto_ano_veiculo.place(x= 403, y= 20)
container_ano_veiculo = ctk.CTkEntry(master= dados_veiculo)
container_ano_veiculo.place(x= 400, y= 45)


# inserir Placa veiculo
texto_placa_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Placa' )
texto_placa_veiculo.place(x= 43, y= 90)
container_placa_veiculo = ctk.CTkEntry(master= dados_veiculo)
container_placa_veiculo.place(x= 40, y= 115)

# inserir Chassi veiculo
texto_chassi_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Chassi' )
texto_chassi_veiculo.place(x= 193, y= 90)
container_chassi_veiculo = ctk.CTkEntry(master= dados_veiculo)
container_chassi_veiculo.place(x= 190, y= 115)

# inserir renavan veiculo
texto_renavan_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Renavan' )
texto_renavan_veiculo.place(x= 343, y= 90)
container_renavan_veiculo = ctk.CTkEntry(master= dados_veiculo)
container_renavan_veiculo.place(x= 340, y= 115)

def limpar_conteudo():
    container_codigo_veiculo.configure(state='normal')
    container_codigo_veiculo.delete(0,'end')
    container_marca_veiculo.configure(state='normal')
    container_marca_veiculo.delete(0,'end')
    container_modelo_veiculo.configure(state='normal')
    container_modelo_veiculo.delete(0,'end')
    container_ano_veiculo.configure(state='normal')
    container_ano_veiculo.delete(0,'end')
    container_placa_veiculo.configure(state='normal')
    container_placa_veiculo.delete(0,'end')
    container_chassi_veiculo.configure(state='normal')
    container_chassi_veiculo.delete(0,'end')
    container_renavan_veiculo.configure(state='normal')
    container_renavan_veiculo.delete(0,'end')
    
botao_limpar = ctk.CTkButton(dados_veiculo, text='Limpar', fg_color='light yellow', font=('Arial', 10), text_color='black', width=30, hover_color='grey', command=limpar_conteudo)
botao_limpar.place(x=530, y=4)

#################################################################### BOTÃO CADASTRAR VEICULOS ############################################################################################################
def cadastrar_veiculo():
    app_5 = ctk.CTkToplevel(app_2)
    app_5.lift()
    app_5.focus_force()
    app_5.grab_set()
    app_5.title('Cadastrar Veículo')
    app_5.maxsize(width=800, height=300)
    codigo_veiculo = gerar_codigo_veiculo()
    # frame de dados do veiculo 
    dados_veiculo = ctk.CTkFrame(app_5, width=580, height=200, border_width=2)
    dados_veiculo.pack(pady= 10)
    # codigo veiculo
    texto_codigo_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Código' )
    texto_codigo_veiculo.place(x= 43, y= 20)
    container_codigo_veiculo = ctk.CTkEntry(master= dados_veiculo, width=50)
    container_codigo_veiculo.insert(0, codigo_veiculo)
    container_codigo_veiculo.configure(state='readonly', text_color= 'grey')
    container_codigo_veiculo.place(x= 40, y= 45)
    

    # inserir marca do veiculo
    texto_marca_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Marca' )
    texto_marca_veiculo.place(x= 103, y= 20)
    container_marca_veiculo = ctk.CTkEntry(master= dados_veiculo, fg_color='white', text_color='black')
    container_marca_veiculo.place(x= 100, y= 45)
    

    # inserir modelo do veiculo
    texto_modelo_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Modelo' )
    texto_modelo_veiculo.place(x= 253, y= 20)
    container_modelo_veiculo = ctk.CTkEntry(master= dados_veiculo,fg_color='white', text_color='black')
    container_modelo_veiculo.place(x= 250, y= 45)
    
    

    # inserir ano veiculo
    texto_ano_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Ano' )
    texto_ano_veiculo.place(x= 403, y= 20)
    container_ano_veiculo = ctk.CTkEntry(master= dados_veiculo,fg_color='white', text_color='black')
    container_ano_veiculo.place(x= 400, y= 45)
    


    # inserir Placa veiculo
    texto_placa_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Placa' )
    texto_placa_veiculo.place(x= 43, y= 90)
    container_placa_veiculo = ctk.CTkEntry(master= dados_veiculo,fg_color='white', text_color='black')
    container_placa_veiculo.place(x= 40, y= 115)
    

    # inserir Chassi veiculo
    texto_chassi_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'Chassi' )
    texto_chassi_veiculo.place(x= 193, y= 90)
    container_chassi_veiculo = ctk.CTkEntry(master= dados_veiculo,fg_color='white', text_color='black')
    container_chassi_veiculo.place(x= 190, y= 115)
    

    # inserir renavan veiculo
    texto_renavan_veiculo = ctk.CTkLabel(master= dados_veiculo, text= 'renavan' )
    texto_renavan_veiculo.place(x= 343, y= 90)
    container_renavan_veiculo = ctk.CTkEntry(master= dados_veiculo,fg_color='white', text_color='black')
    container_renavan_veiculo.place(x= 340, y= 115)
    
    texto_confirmacao = ctk.CTkLabel(master= dados_veiculo, text= '')
    texto_confirmacao.place(x= 100, y=170)
    
    
    # confirmar
    def botao_confirmar_app5():
        if len(container_placa_veiculo.get()) < 7:
            texto_confirmacao.configure(text= 'Digitos da placa insuficientes', text_color = 'red')
        elif container_marca_veiculo.get() == '' or None:
            texto_confirmacao.configure(text= 'Insira a marca', text_color = 'red')
        elif container_modelo_veiculo.get() =='' or None:
            texto_confirmacao.configure(text= 'Insira o modelo', text_color = 'red')
        else:
            try:
                cursor_vei.execute('''
                    INSERT INTO VEICULOS (
                        "CÓDIGO", "MARCA", "MODELO", "ANO", "PLACA", "CHASSI", "renavan"
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    container_codigo_veiculo.get().upper(),
                    container_marca_veiculo.get().upper(),
                    container_modelo_veiculo.get().upper(),
                    container_ano_veiculo.get().upper(),
                    container_placa_veiculo.get().upper(),
                    container_chassi_veiculo.get().upper(),
                    container_renavan_veiculo.get().upper()
                ))

                banco_vei.commit() 
                
                texto_confirmacao.configure(text= 'Cadastro Confirmado', text_color = 'green')
                
                container_codigo_veiculo.delete(0,'end')
                container_marca_veiculo.delete(0,'end')
                container_modelo_veiculo.delete(0,'end')
                container_ano_veiculo.delete(0,'end')
                container_placa_veiculo.delete(0,'end')
                container_chassi_veiculo.delete(0,'end')
                container_renavan_veiculo.delete(0,'end')
                
                novo_codigo = gerar_codigo_veiculo()
                
                container_codigo_veiculo.configure(state='normal')
                container_codigo_veiculo.delete(0,'end')
                container_codigo_veiculo.insert(0, novo_codigo)
                container_codigo_veiculo.configure(state='readonly')
                
                
                    
                
            except Exception as erro:
                texto_confirmacao.configure(master= dados_veiculo, text= f'ERRO: {erro}', text_color = 'red')
             
        
    botao_confirmar_veiculo = ctk.CTkButton(app_5, fg_color='green', image=simbolo_confirmar, text= '', hover_color='dark green', width=80, command=botao_confirmar_app5)
    botao_confirmar_veiculo.place(x=350, y=170 )
    # cancelar
    def botao_cancelar_app5():
        app_5.destroy()
    botao_cancelar_veiculo = ctk.CTkButton(app_5, fg_color='red', image=simbolo_cancelar, text= '', hover_color='dark red', width=80, command=botao_cancelar_app5)
    botao_cancelar_veiculo.place(x=450, y=170 )
    
    
botao_cadastrar_veiculo = ctk.CTkButton(dados_veiculo,image= simbolo_adicionar, text='', height=25,width=80, font=('Arial', 12, 'bold'), fg_color='green', hover_color='dark green', command=cadastrar_veiculo)
botao_cadastrar_veiculo.place(x=350, y=160 )

######################################################################### botao procurar veiculo #################################################################################################################################
def exibir_tabela_vei():
    app_8 = ctk.CTkToplevel(app_2)
    app_8.geometry('1000x220')
    app_8.title('TABELA VEICULOS')
    app_8.lift()
    app_8.focus_force()
    app_8.grab_set()
    
    # Frame principal
    frame_tabela = ctk.CTkFrame(app_8)
    frame_tabela.pack(fill='both', expand=True)
    
    tabela_vei = ttk.Treeview(frame_tabela,columns=(
    "CÓDIGO",
    "MARCA",
    "MODELO",
    "ANO",
    "PLACA",
    "CHASSI",
    "RENAVAN"
    ),
    show='headings'
    )
    
    # Configura colunas e cabeçalhos
    for col in tabela_vei["columns"]:
        tabela_vei.heading(col, text=col)
        tabela_vei.column(col, width=150, anchor='center')

    tabela_vei.column("CÓDIGO", width=80)

    # Scrollbars
    scroll_y = ttk.Scrollbar(frame_tabela, orient='vertical', command=tabela_vei.yview)
    scroll_x = ttk.Scrollbar(frame_tabela, orient='horizontal', command=tabela_vei.xview)

    tabela_vei.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

    # Posicionamento com grid
    tabela_vei.grid(row=0, column=0, sticky='nsew')
    scroll_y.grid(row=0, column=1, sticky='ns')
    scroll_x.grid(row=1, column=0, sticky='ew')

    # Expansão automática
    frame_tabela.grid_rowconfigure(0, weight=1)
    frame_tabela.grid_columnconfigure(0, weight=1)

    # Exibe dados se campos nome e cpf estiverem vazios
    if not container_renavan_veiculo.get() and not container_placa_veiculo.get():
        conexao = sqlite3.connect(caminho_recurso('VEICULOS.db'))
        cursor = conexao.cursor()
        cursor.execute('SELECT * FROM VEICULOS')
        dados = cursor.fetchall()
        
        for linha in dados:
            tabela_vei.insert('', 'end', values=linha)
            conexao.close()
            
    elif len(container_placa_veiculo.get()) > 1:
        conexao = sqlite3.connect(caminho_recurso('VEICULOS.db'))
        cursor = conexao.cursor()
        container_placa_veiculo.get()
        cursor.execute(f'SELECT * FROM VEICULOS WHERE PLACA LIKE "%{container_placa_veiculo.get()}%"')
        dados = cursor.fetchall()
        for linha in dados:
            tabela_vei.insert('', 'end', values=linha)
        
            
            
            
    elif len(container_renavan_veiculo.get()) > 1:
        conexao = sqlite3.connect(caminho_recurso('VEICULOS.db'))
        cursor = conexao.cursor()
        container_renavan_veiculo.get()
        cursor.execute(f'SELECT * FROM VEICULOS WHERE NOME LIKE "%{container_renavan_veiculo.get()}%"')
        dados = cursor.fetchall()
        for linha in dados:
            tabela_vei.insert('', 'end', values=linha)


    
    def adicionar_dados(event):
        selecionado = tabela_vei.focus()
        if selecionado:
            valores = tabela_vei.item(selecionado, 'values')
            if valores:
                container_codigo_veiculo.delete(0, 'end')
                container_codigo_veiculo.insert(0, valores[0] if len(valores) > 0 else '')
                container_codigo_veiculo.configure(state= 'readonly', text_color= 'grey')
                container_marca_veiculo.delete(0, 'end')
                container_marca_veiculo.insert(0, valores[1] if len(valores) > 0 else '')
                container_marca_veiculo.configure(state= 'readonly', text_color= 'grey')
                container_modelo_veiculo.delete(0, 'end')
                container_modelo_veiculo.insert(0, valores[2] if len(valores) > 0 else '')
                container_modelo_veiculo.configure(state= 'readonly', text_color= 'grey')
                container_ano_veiculo.delete(0, 'end')
                container_ano_veiculo.insert(0, valores[3] if len(valores) > 0 else '')
                container_ano_veiculo.configure(state= 'readonly', text_color= 'grey')
                container_placa_veiculo.delete(0, 'end')
                container_placa_veiculo.insert(0, valores[4] if len(valores) > 0 else '')
                container_placa_veiculo.configure(state= 'readonly', text_color= 'grey')
                container_chassi_veiculo.delete(0, 'end')
                container_chassi_veiculo.insert(0, valores[5] if len(valores) > 0 else '')
                container_chassi_veiculo.configure(state= 'readonly', text_color= 'grey')
                container_renavan_veiculo.delete(0, 'end')
                container_renavan_veiculo.insert(0, valores[6] if len(valores) > 0 else '')
                container_renavan_veiculo.configure(state= 'readonly', text_color= 'grey')
    
    tabela_vei.bind('<Double-1>',adicionar_dados )


botao_procurar_veiculo = ctk.CTkButton(dados_veiculo, image=simbolo_procurar, text='', height=25, width=80, font=('Arial', 12, 'bold'), command=exibir_tabela_vei)
botao_procurar_veiculo.place(x=450, y=160 )
################################################################################################################################################################################################################################################

########################## Dados do cliente ##############################

# Titulo bloco dados cliente
titulo_frame_3 =  ctk.CTkLabel(frame_titulos, font=('Arial', 14, 'bold'), text='Dados Cliente')
titulo_frame_3.place(x=800, y=0)

# Frame dados cliente
dados_cliente = ctk.CTkFrame(frame_veiculo_extras, width=580, height=200, border_width=2 )
dados_cliente.pack(padx=10, side= 'left')


# codigo cliente
texto_codigo_cliente = ctk.CTkLabel(master= dados_cliente, text= 'Código' )
texto_codigo_cliente.place(x= 43, y= 20)
container_codigo_cliente = ctk.CTkEntry(master= dados_cliente, width=50)
container_codigo_cliente.place(x= 40, y= 45)

# inserir nome cliente
texto_nome_cliente = ctk.CTkLabel(master= dados_cliente, text= 'Nome' )
texto_nome_cliente.place(x= 103, y= 20)
container_nome_cliente = ctk.CTkEntry(master= dados_cliente)
container_nome_cliente.place(x= 100, y= 45)

# cnpj cliente
def formatar_cnpj(event=None):
    texto = container_cnpj_cliente.get()
    texto = re.sub(r'\D', '', texto)
    texto_limpo = ''.join(filter(str.isdigit, texto))

    texto_formatado = ''
    if len(texto_limpo) >= 1:
        texto_formatado += texto_limpo[:2]
    if len(texto_limpo) >= 3:
        texto_formatado = texto_formatado[:2] + '.' + texto_limpo[2:5]
    if len(texto_limpo) >= 6:
        texto_formatado = texto_formatado[:6] + '.' + texto_limpo[5:8]
    if len(texto_limpo) >= 9:
        texto_formatado = texto_formatado[:10] + '/' + texto_limpo[8:12]
    if len(texto_limpo) >= 13:
        texto_formatado = texto_formatado[:15] + '-' + texto_limpo[12:14]

    container_cnpj_cliente.delete(0, 'end')
    container_cnpj_cliente.insert(0, texto_formatado)
    
texto_cnpj_cliente = ctk.CTkLabel(master= dados_cliente, text= 'CNPJ' )
texto_cnpj_cliente.place(x= 253, y= 20)
container_cnpj_cliente = ctk.CTkEntry(master=dados_cliente)
container_cnpj_cliente.place(x= 250, y= 45)
container_cnpj_cliente.bind("<KeyRelease>", formatar_cnpj)


# inserir inscrição estadual
def formatar_insc_estadual(event=None):
    texto = container_insc_estadual.get()

    # Remove tudo que não for número
    numeros = ''.join(filter(str.isdigit, texto))

    # Insere o ponto a cada 3 números
    partes = [numeros[i:i+3] for i in range(0, len(numeros), 3)]
    texto_formatado = '.'.join(partes)

    # Atualiza o campo com o texto formatado
    container_insc_estadual.delete(0, 'end')
    container_insc_estadual.insert(0, texto_formatado)
texto_insc_estadual = ctk.CTkLabel(master= dados_cliente, text= 'Insc. Estadual' )
texto_insc_estadual.place(x= 403, y= 20)
container_insc_estadual = ctk.CTkEntry(master= dados_cliente)
container_insc_estadual.place(x= 400, y= 45)
container_insc_estadual.bind('<KeyRelease>', formatar_insc_estadual)



# inserir Endereço cliente
texto_endereco_cliente = ctk.CTkLabel(master= dados_cliente, text= 'Endereço' )
texto_endereco_cliente.place(x= 43, y= 90)
container_endereco_cliente = ctk.CTkEntry(master= dados_cliente)
container_endereco_cliente.place(x= 40, y= 115)

# inserir municipio cliente
texto_municipio_cliente = ctk.CTkLabel(master= dados_cliente, text= 'Município' )
texto_municipio_cliente.place(x= 193, y= 90)
container_municipio_cliente = ctk.CTkEntry(master= dados_cliente)
container_municipio_cliente.place(x= 190, y= 115)




def limpar_conteudo():
    container_codigo_cliente.configure(state='normal')
    container_codigo_cliente.delete(0,'end')
    container_nome_cliente.configure(state='normal')
    container_nome_cliente.delete(0,'end')
    container_cnpj_cliente.configure(state='normal')
    container_cnpj_cliente.delete(0,'end')
    container_endereco_cliente.configure(state='normal')
    container_endereco_cliente.delete(0,'end')
    container_municipio_cliente.configure(state='normal')
    container_municipio_cliente.delete(0,'end')

    
botao_limpar = ctk.CTkButton(dados_cliente, text='Limpar', fg_color='light yellow', font=('Arial', 10), text_color='black', width=30, hover_color='grey', command=limpar_conteudo)
botao_limpar.place(x=530, y=4)

#################################################################### BOTÃO CADASTRAR CLIENTES ############################################################################################################
def cadastrar_cliente():
    app_9 = ctk.CTkToplevel(app_2)
    app_9.lift()
    app_9.focus_force()
    app_9.grab_set()
    app_9.title('Cadastrar Clientes')
    app_9.maxsize(width=800, height=300)
    codigo_cliente = gerar_codigo_cliente()
    # frame de dados do cliente 
    dados_cliente = ctk.CTkFrame(app_9, width=580, height=300, border_width=2)
    dados_cliente.pack(pady= 10)
    # codigo cliente
    texto_codigo_cliente = ctk.CTkLabel(master= dados_cliente, text= 'Código' )
    texto_codigo_cliente.place(x= 43, y= 20)
    container_codigo_cliente = ctk.CTkEntry(master= dados_cliente, width=50)
    container_codigo_cliente.insert(0, codigo_cliente)
    container_codigo_cliente.configure(state='readonly', text_color= 'grey')
    container_codigo_cliente.place(x= 40, y= 45)
    

    # inserir nome cliente
    texto_nome_cliente = ctk.CTkLabel(master= dados_cliente, text= 'Nome' )
    texto_nome_cliente.place(x= 103, y= 20)
    container_nome_cliente = ctk.CTkEntry(master= dados_cliente)
    container_nome_cliente.place(x= 100, y= 45)

    # cnpj cliente
    def formatar_cnpj(event=None):
        texto = container_cnpj_cliente.get()
        texto = re.sub(r'\D', '', texto)
        texto_limpo = ''.join(filter(str.isdigit, texto))

        texto_formatado = ''
        if len(texto_limpo) >= 1:
            texto_formatado += texto_limpo[:2]
        if len(texto_limpo) >= 3:
            texto_formatado = texto_formatado[:2] + '.' + texto_limpo[2:5]
        if len(texto_limpo) >= 6:
            texto_formatado = texto_formatado[:6] + '.' + texto_limpo[5:8]
        if len(texto_limpo) >= 9:
            texto_formatado = texto_formatado[:10] + '/' + texto_limpo[8:12]
        if len(texto_limpo) >= 13:
            texto_formatado = texto_formatado[:15] + '-' + texto_limpo[12:14]

        container_cnpj_cliente.delete(0, 'end')
        container_cnpj_cliente.insert(0, texto_formatado)
        
    texto_cnpj_cliente = ctk.CTkLabel(master= dados_cliente, text= 'CNPJ' )
    texto_cnpj_cliente.place(x= 253, y= 20)
    container_cnpj_cliente = ctk.CTkEntry(master=dados_cliente)
    container_cnpj_cliente.place(x= 250, y= 45)
    container_cnpj_cliente.bind("<KeyRelease>", formatar_cnpj)


    # inserir inscrição estadual
    def formatar_insc_estadual(event=None):
        texto = container_insc_estadual.get()

        # Remove tudo que não for número
        numeros = ''.join(filter(str.isdigit, texto))

        # Insere o ponto a cada 3 números
        partes = [numeros[i:i+3] for i in range(0, len(numeros), 3)]
        texto_formatado = '.'.join(partes)

        # Atualiza o campo com o texto formatado
        container_insc_estadual.delete(0, 'end')
        container_insc_estadual.insert(0, texto_formatado)
    texto_insc_estadual = ctk.CTkLabel(master= dados_cliente, text= 'Insc. Estadual' )
    texto_insc_estadual.place(x= 403, y= 20)
    container_insc_estadual = ctk.CTkEntry(master= dados_cliente)
    container_insc_estadual.place(x= 400, y= 45)
    container_insc_estadual.bind('<KeyRelease>', formatar_insc_estadual)



    # inserir Endereço cliente
    texto_endereco_cliente = ctk.CTkLabel(master= dados_cliente, text= 'Endereço' )
    texto_endereco_cliente.place(x= 43, y= 90)
    container_endereco_cliente = ctk.CTkEntry(master= dados_cliente)
    container_endereco_cliente.place(x= 40, y= 115)

    # inserir municipio cliente
    texto_municipio_cliente = ctk.CTkLabel(master= dados_cliente, text= 'Município' )
    texto_municipio_cliente.place(x= 193, y= 90)
    container_municipio_cliente = ctk.CTkEntry(master= dados_cliente)
    container_municipio_cliente.place(x= 190, y= 115)
    
    texto_confirmacao = ctk.CTkLabel(master= dados_cliente, text= '')
    texto_confirmacao.place(x= 100, y=170)
    
    # confirmar
    def botao_confirmar_app9():
        if container_nome_cliente.get() == '' or None:
            texto_confirmacao.configure(text= 'Insira o nome', text_color = 'red')
        elif len(container_cnpj_cliente.get()) < 18: 
            texto_confirmacao.configure(text= 'CNPJ inválido', text_color = 'red')
        else:
            try:
                cursor_cli.execute('''
                    INSERT INTO CLIENTES (
                        "CÓDIGO", "NOME", "CNPJ", "INSC. ESTADUAL", "ENDEREÇO", "MUNICIPIO"
                    )
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (
                    container_codigo_cliente.get().upper(),
                    container_nome_cliente.get().upper(),
                    container_cnpj_cliente.get().upper(),
                    container_insc_estadual.get().upper(),
                    container_endereco_cliente.get().upper(),
                    container_municipio_cliente.get().upper(),
                ))

                banco_cli.commit() 
                
                texto_confirmacao.configure(text= 'Cadastro Confirmado', text_color = 'green')
                
                container_codigo_cliente.delete(0,'end')
                container_nome_cliente.delete(0,'end')
                container_cnpj_cliente.delete(0,'end')
                container_insc_estadual.delete(0,'end')
                container_endereco_cliente.delete(0,'end')
                container_municipio_cliente.delete(0,'end')
                
                novo_codigo = gerar_codigo_cliente()
                
                container_codigo_cliente.configure(state='normal')
                container_codigo_cliente.delete(0,'end')
                container_codigo_cliente.insert(0, novo_codigo)
                container_codigo_cliente.configure(state='readonly')
                       
            except Exception as erro:
                texto_confirmacao.configure(master= dados_cliente, text= f'ERRO: {erro}', text_color = 'red')
             
        
    botao_confirmar_cliente = ctk.CTkButton(app_9, fg_color='green', image=simbolo_confirmar, text= '', hover_color='dark green', width=80, command=botao_confirmar_app9)
    botao_confirmar_cliente.place(x=350, y=170 )
    # cancelar
    def botao_cancelar_app9():
        app_9.destroy()
    botao_cancelar_cliente = ctk.CTkButton(app_9, fg_color='red', image=simbolo_cancelar, text= '', hover_color='dark red', width=80, command=botao_cancelar_app9)
    botao_cancelar_cliente.place(x=450, y=170 )
    
    
botao_cadastrar_cliente = ctk.CTkButton(dados_cliente,image= simbolo_adicionar, text='', height=25,width=80, font=('Arial', 12, 'bold'), fg_color='green', hover_color='dark green', command=cadastrar_cliente)
botao_cadastrar_cliente.place(x=350, y=160 )

######################################################################### botao procurar cliente #################################################################################################################################
def exibir_tabela_cli():
    app_10 = ctk.CTkToplevel(app_2)
    app_10.geometry('1000x220')
    app_10.title('TABELA CLIENTES')
    app_10.lift()
    app_10.focus_force()
    app_10.grab_set()
    
    # Frame principal
    frame_tabela = ctk.CTkFrame(app_10)
    frame_tabela.pack(fill='both', expand=True)
    
    tabela_cli = ttk.Treeview(frame_tabela,columns=(
    "CÓDIGO",
    "NOME",
    "CNPJ",
    "INSC. ESTADUAL",
    "ENDEREÇO",
    "MUNICIPIO",
    ),
    show='headings'
    )
    
    # Configura colunas e cabeçalhos
    for col in tabela_cli["columns"]:
        tabela_cli.heading(col, text=col)
        tabela_cli.column(col, width=150, anchor='center')

    tabela_cli.column("CÓDIGO", width=80)

    # Scrollbars
    scroll_y = ttk.Scrollbar(frame_tabela, orient='vertical', command=tabela_cli.yview)
    scroll_x = ttk.Scrollbar(frame_tabela, orient='horizontal', command=tabela_cli.xview)

    tabela_cli.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

    # Posicionamento com grid
    tabela_cli.grid(row=0, column=0, sticky='nsew')
    scroll_y.grid(row=0, column=1, sticky='ns')
    scroll_x.grid(row=1, column=0, sticky='ew')

    # Expansão automática
    frame_tabela.grid_rowconfigure(0, weight=1)
    frame_tabela.grid_columnconfigure(0, weight=1)

    # Exibe dados se campos nome e cpf estiverem vazios
    if not container_nome_cliente.get() and not container_cnpj_cliente.get():
        conexao = sqlite3.connect(caminho_recurso('CLIENTES.db'))
        cursor = conexao.cursor()
        cursor.execute('SELECT * FROM CLIENTES')
        dados = cursor.fetchall()
        
        for linha in dados:
            tabela_cli.insert('', 'end', values=linha)
            conexao.close()
            
    elif len(container_nome_cliente.get()) > 1:
        conexao = sqlite3.connect(caminho_recurso('CLIENTES.db'))
        cursor = conexao.cursor()
        container_nome_cliente.get()
        cursor.execute(f'SELECT * FROM CLIENTES WHERE NOME LIKE "%{container_nome_cliente.get()}%"')
        dados = cursor.fetchall()
        for linha in dados:
            tabela_cli.insert('', 'end', values=linha)
        
            
            
            
    elif len(container_cnpj_cliente.get()) > 1:
        conexao = sqlite3.connect(caminho_recurso('CLIENTES.db'))
        cursor = conexao.cursor()
        container_cnpj_cliente.get()
        cursor.execute(f'SELECT * FROM CLIENTES WHERE CNPJ LIKE "%{container_cnpj_cliente.get()}%"')
        dados = cursor.fetchall()
        for linha in dados:
            tabela_cli.insert('', 'end', values=linha)


    
    def adicionar_dados(event):
        selecionado = tabela_cli.focus()
        if selecionado:
            valores = tabela_cli.item(selecionado, 'values')
            if valores:
                container_codigo_cliente.delete(0, 'end')
                container_codigo_cliente.insert(0, valores[0] if len(valores) > 0 else '')
                container_codigo_cliente.configure(state= 'readonly', text_color= 'grey')
                container_nome_cliente.delete(0, 'end')
                container_nome_cliente.insert(0, valores[1] if len(valores) > 0 else '')
                container_nome_cliente.configure(state= 'readonly', text_color= 'grey')
                container_cnpj_cliente.delete(0, 'end')
                container_cnpj_cliente.insert(0, valores[2] if len(valores) > 0 else '')
                container_cnpj_cliente.configure(state= 'readonly', text_color= 'grey')
                container_insc_estadual.delete(0, 'end')
                container_insc_estadual.insert(0, valores[3] if len(valores) > 0 else '')
                container_insc_estadual.configure(state= 'readonly', text_color= 'grey')
                container_endereco_cliente.delete(0, 'end')
                container_endereco_cliente.insert(0, valores[4] if len(valores) > 0 else '')
                container_endereco_cliente.configure(state= 'readonly', text_color= 'grey')
                container_municipio_cliente.delete(0, 'end')
                container_municipio_cliente.insert(0, valores[5] if len(valores) > 0 else '')
                container_municipio_cliente.configure(state= 'readonly', text_color= 'grey')
    
    tabela_cli.bind('<Double-1>',adicionar_dados )


botao_procurar_cliente = ctk.CTkButton(dados_cliente, image=simbolo_procurar, text='', height=25, width=80, font=('Arial', 12, 'bold'), command=exibir_tabela_cli)
botao_procurar_cliente.place(x=450, y=160 )

########################## Extras ##############################

# Titulo bloco dados extras
titulo_frame_3 =  ctk.CTkLabel(frame_titulos, font=('Arial', 14, 'bold'), text='Dados Extras')
titulo_frame_3.place(x= 1305, y=0)

# Frame dados extras
dados_extras = ctk.CTkFrame(frame_veiculo_extras, width=400, height=250, border_width=2 )
dados_extras.pack(padx= 10, side= 'left')


# inserir peso veiculo
def formatar_milhar(event):
    entry = event.widget
    texto = entry.get()

    # Se acabou de digitar a vírgula, não formata ainda
    if texto.endswith(','):
        return

    # Quebra em parte inteira e decimal, se tiver
    if ',' in texto:
        parte_inteira, parte_decimal = texto.split(',', 1)
    else:
        parte_inteira, parte_decimal = texto, ''

    # Remove pontos da parte inteira
    parte_inteira = parte_inteira.replace('.', '')

    # Se não for só número, ignora
    if not parte_inteira.isdigit():
        return

    # Formata com ponto de milhar
    parte_inteira_formatada = f"{int(parte_inteira):,}".replace(',', '.')

    # Monta de volta
    if parte_decimal:
        texto_formatado = f"{parte_inteira_formatada},{parte_decimal}"
    else:
        texto_formatado = parte_inteira_formatada

    # Atualiza o campo sem apagar vírgula
    entry.delete(0, 'end')
    entry.insert(0, texto_formatado)
    
texto_peso_veiculo = ctk.CTkLabel(master= dados_extras, text= 'Peso' )
texto_peso_veiculo.place(x= 83, y= 4)
container_peso_veiculo = ctk.CTkEntry(master= dados_extras)
container_peso_veiculo.place(x= 30, y= 25)
container_peso_veiculo.bind("<KeyRelease>", formatar_milhar)


# inserir frete
texto_frete = ctk.CTkLabel(master= dados_extras, text= 'Frete/ton' )
texto_frete.place(x= 75, y= 55)
container_frete = ctk.CTkEntry(master= dados_extras)
container_frete.place(x= 30, y= 75)
container_frete.bind("<KeyRelease>", formatar_milhar)


# inserir adiantamento
texto_adiantamento = ctk.CTkLabel(master= dados_extras, text= 'Adiantamento' )
texto_adiantamento.place(x= 62, y= 138)
container_adiantamento = ctk.CTkEntry(master= dados_extras)
container_adiantamento.place(x= 30, y= 160)
container_adiantamento.bind("<KeyRelease>", formatar_milhar)

# dados da carga 
texto_dados_carga = ctk.CTkLabel(master= dados_extras, text= 'Dados da Carga' )
texto_dados_carga.place(x= 243, y= 4)
container_dados_carga = ctk.CTkEntry(master= dados_extras)
container_dados_carga.place(x= 220, y= 25)
container_dados_carga.bind("<KeyRelease>")


# inserir coleta
texto_coleta = ctk.CTkLabel(master= dados_extras, text= 'Coleta' )
texto_coleta.place(x= 268, y= 55)
container_coleta = ctk.CTkEntry(master= dados_extras)
container_coleta.place(x= 220, y= 75)

# inserir NF

def formatar_nf(event=None):
        # Remove tudo que não for número
        texto = container_nf.get()
        texto = re.sub(r'\D', '', texto)
        container_nf.delete(0, 'end')
        container_nf.insert(0, texto)
        
        texto = container_nf.get()

        # Remove tudo que não for número
        numeros = ''.join(filter(str.isdigit, texto))

        # Insere o ponto a cada 3 números
        partes = [numeros[i:i+3] for i in range(0, len(numeros), 3)]
        texto_formatado = '.'.join(partes)

        # Atualiza o campo com o texto formatado
        container_nf.delete(0, 'end')
        container_nf.insert(0, texto_formatado)
        
texto_nf = ctk.CTkLabel(master= dados_extras, text= 'NF' )
texto_nf.place(x= 283, y= 105)
container_nf = ctk.CTkEntry(master= dados_extras)
container_nf.place(x= 220, y= 125)
container_nf.bind("<KeyRelease>", formatar_milhar)

# inserir data de lançamento
def formatar_data_lancamento(event=None):
        texto = container_data_lancamento.get()
        texto_limpo = ''.join(filter(str.isdigit, texto))[:8]  # Máximo 8 dígitos

        texto_formatado = ''
        if len(texto_limpo) >= 1:
            texto_formatado += texto_limpo[:2]
        if len(texto_limpo) >= 3:
            texto_formatado = texto_formatado[:2] + '/' + texto_limpo[2:4]
        if len(texto_limpo) >= 5:
            texto_formatado = texto_formatado[:5] + '/' + texto_limpo[4:8]

        container_data_lancamento.delete(0, 'end')
        container_data_lancamento.insert(0, texto_formatado)
texto_data_lancamento = ctk.CTkLabel(master= dados_extras, text= 'Data de saída' )
texto_data_lancamento.place(x= 248, y= 155)
container_data_lancamento = ctk.CTkEntry(master= dados_extras)
container_data_lancamento.place(x= 220, y= 175)
container_data_lancamento.bind("<KeyRelease>", formatar_data_lancamento)

def formatar_moeda(valor):
    return f'{valor:,.2f}'.replace(',', 'v').replace('.', ',').replace('v', '.')
# texto total e saldo
texto_total = ctk.CTkLabel(dados_extras, text='Total =', text_color='grey', font=('Arial', 12, 'bold'))
texto_total.place(x=40, y=105)
texto_saldo = ctk.CTkLabel(dados_extras, text='Saldo =', text_color='grey', font=('Arial', 12, 'bold'))
texto_saldo.place(x=40,y=190)


def pegar_valor_formatado(entry):
    texto = entry.get()
    texto = texto.replace('.', '').replace(',', '.')
    try:
        return float(texto)
    except ValueError:
        return 0.0

def calcular():
    peso = pegar_valor_formatado(container_peso_veiculo) / 1000
    frete = pegar_valor_formatado(container_frete)
    adiantamento = pegar_valor_formatado(container_adiantamento)
    try:
        total = frete * peso
        total_formatado = formatar_moeda(total)

        texto_total.configure(
            text=f'Total = (R${total_formatado})',
            text_color='green',
            font=('Arial', 12, 'bold')
        )

        if adiantamento != '':
            saldo = total - int(adiantamento)
            saldo_formatado = formatar_moeda(saldo)

            texto_saldo.configure(
                text=f'Saldo = (R${saldo_formatado})',
                text_color='green',
                font=('Arial', 12, 'bold')
            )
        else:
            texto_saldo.configure(
                text='Saldo =',
                text_color='grey',
                font=('Arial', 12, 'bold')
            )
    except ValueError:
        texto_total.configure(
            text='Total = ERRO',
            text_color='red',
            font=('Arial', 12, 'bold')
        )

def confirmar():
    calcular()
    
    container_peso_veiculo.configure(state='readonly', text_color='grey')
    container_frete.configure(state='readonly', text_color='grey') 
    container_adiantamento.configure(state='readonly', text_color='grey') 
    container_dados_carga.configure(state='readonly', text_color='grey') 
    container_coleta.configure(state='readonly', text_color='grey') 
    container_nf.configure(state='readonly', text_color='grey') 
    container_data_lancamento.configure(state='readonly', text_color='grey')  

botao_calcular = ctk.CTkButton(dados_extras, text='', width=20, height=20, fg_color='green', hover_color='dark green', image=simbolo_calcular, command=calcular)
botao_calcular.place(x=280, y=212)

botao_confirmar = ctk.CTkButton(dados_extras, text='',  width=20, height=20, fg_color='green', hover_color='dark green', image=simbolo_confirmar, command=confirmar)
botao_confirmar.place(x=320, y=212)

def limpar_conteudo():

    container_peso_veiculo.configure(state='normal', text_color='white')
    container_peso_veiculo.delete(0, 'end')
    container_frete.configure(state='normal', text_color='white') 
    container_frete.delete(0, 'end')    
    container_adiantamento.configure(state='normal', text_color='white')
    container_adiantamento.delete(0, 'end') 
    container_dados_carga.configure(state='normal', text_color='white')
    container_dados_carga.delete(0, 'end') 
    container_coleta.configure(state='normal', text_color='white')
    container_coleta.delete(0, 'end') 
    container_nf.configure(state='normal', text_color='white')
    container_nf.delete(0, 'end') 
    container_data_lancamento.configure(state='normal', text_color='white')
    container_data_lancamento.delete(0, 'end')  
    texto_total.configure(text= 'Total =')
    texto_saldo.configure(text= 'Saldo =')    


botao_limpar = ctk.CTkButton(dados_extras, text='Limpar', fg_color='light yellow', font=('Arial', 10), text_color='black', width=26, hover_color='grey', command=limpar_conteudo, height=23)
botao_limpar.place(x=225, y=212)

########################## Botâo emitir carta frete ##############################

def emitir():
    numero_recibo = gerar_numero_recibo()
    
    cursor_recibo.execute(f'INSERT INTO NUMERO_RECIBO ("NUMERO DO RECIBO") VALUES(?)', (numero_recibo,))
    banco_recibo.commit()
    
    def limpar_entry(entry):
        texto = entry.get().strip()
        texto = texto.replace('.', '').replace(',', '.')
        try:
            valor = float(texto)
            return valor
        except ValueError:
            return 0.0
        
    frete = limpar_entry(container_frete)    
    frete = float(frete) if frete else 0
    peso = limpar_entry(container_peso_veiculo)
    peso = float(peso/1000) if peso else 0
    adiantamento = limpar_entry(container_adiantamento)
    adiantamento = float(adiantamento) 
    total = frete * peso
    preco_total = formatar_moeda(total)
    try:
        planilha = load_workbook(caminho_recurso('RECIBO DE FRETE.xlsx'))

        sheet = planilha.active 
        
        fonte_corpo = Font(name='Roboto',size=10,bold=True, italic=False, color='000000' )
        fonte_titulo = Font(name='Roboto',size=12,bold=True, italic=False, color='000000' )
        fonte_vermelha = Font(name='Roboto',size=20,bold=True, italic=False, color='FF0000' )
        fonte_recibo = Font(name='Roboto',size=26,bold=True, italic=False, color='000000' )
        
        img_logo = Image(caminho_recurso('imagens/logo_plan.png'))
        img_logo.height = 210
        img_logo.width = 320
        img_assinatura = Image(caminho_recurso('imagens/assinatura.png'))
        img_assinatura.height = 90
        img_assinatura.width = 190
        
        sheet.add_image(img_logo, 'A3')
        sheet.add_image(img_assinatura, 'B33')
        
        celula_e6 = sheet['E6'] 
        celula_a11 = sheet['A11'] 
        celula_b11 = sheet['B11'] 
        celula_d11 = sheet['D11'] 
        celula_a14 = sheet['A14'] 
        celula_b14 = sheet['B14'] 
        celula_d14 = sheet['D14'] 
        celula_a18 = sheet['A18'] 
        celula_a19 = sheet['A19'] 
        celula_a20 = sheet['A20'] 
        celula_a21 = sheet['A21'] 
        celula_a22 = sheet['A22'] 
        celula_d18 = sheet['D18'] 
        celula_d19 = sheet['D19'] 
        celula_d20 = sheet['D20'] 
        celula_d21 = sheet['D21'] 
        celula_d22 = sheet['D22'] 
        celula_a27 = sheet['A27']
        celula_b27 = sheet['B27'] 
        celula_c27 = sheet['C27']
        celula_d27 = sheet['D27']
        celula_e27 = sheet['E27']
        celula_f27 = sheet['F27'] 
        celula_f29 = sheet['F29'] 
        celula_f30 = sheet['F30'] 
        celula_e32 = sheet['E32']
        celula_a34 = sheet['B34']
        celula_b4 = sheet['B4']
        celula_b5 = sheet['B5']
        celula_b6 = sheet['B6']
        celula_a8 = sheet['A8']
        celula_a10 = sheet['A10']
        celula_a13 = sheet['A13']
        celula_a17 = sheet['A17']
        celula_a26 = sheet['A26']
        celula_b10 = sheet['B10']
        celula_b13 = sheet['B13']
        celula_b26 = sheet['B26']
        celula_b34 = sheet['B34']
        celula_c26 = sheet['C26']
        celula_d10 = sheet['D10']
        celula_d13 = sheet['D13']
        celula_d17 = sheet['D17']
        celula_d26 = sheet['D26']
        celula_e6 = sheet['E6']
        celula_e26 = sheet['E26']
        celula_e29 = sheet['E29']
        celula_e30 = sheet['E30']
        celula_e32 = sheet['E32']
        celula_f26 = sheet['F30']
        
        
        sheet['E6'] = f'RECIBO Nº {numero_recibo}'
        sheet['A11'] = container_nome_cliente.get()
        sheet['B11'] = container_cnpj_cliente.get()
        sheet['D11'] = container_insc_estadual.get()
        sheet['A14'] = container_data_lancamento.get()
        sheet['B14'] = container_endereco_cliente.get()
        sheet['D14'] = container_municipio_cliente.get()
        sheet['A18'] = f'NOME: {container_nome_motorista.get()}'
        sheet['A19'] = f'RUA: {container_rua_casa_motorista.get()}'
        sheet['A20'] = f'CIDADE: {container_cidade_motorista.get()}'
        sheet['A21'] = f'FONE: {container_telefone_motorista.get()}'
        sheet['A22'] = f'CPF: {container_cpf_motorista.get()}'
        sheet['D18'] = f'PROPRIETÁRIO: {container_nome_proprietario.get()}'
        sheet['D19'] = f'PLACA DO VEÍCULO: {container_placa_veiculo.get()}'
        sheet['D20'] = f'MARCA/MODELO: {container_marca_veiculo.get()} / {container_modelo_veiculo.get()}'
        sheet['D21'] = f'CIDADE PROPRIETÁRIO: {container_cidade_proprietario.get()}'
        sheet['D22'] = f'FONE PROPRIETÁRIO: {container_telefone_proprietario.get()}'
        sheet['A27'] = container_dados_carga.get().upper()
        sheet['B27'] = container_coleta.get().upper()
        sheet['C27'] = container_nf.get()
        sheet['D27'] = container_peso_veiculo.get()
        sheet['E27'] = container_frete.get()
        sheet['F27'] = preco_total
        sheet['F29'] = preco_total
        sheet['F30'] = container_adiantamento.get()
        sheet['E32'] = (total - adiantamento)
        sheet['A34'] = container_nome_motorista.get()
        
        celula_e6.font = fonte_vermelha
        celula_a11.font = fonte_corpo
        celula_b11.font = fonte_corpo
        celula_d11.font = fonte_corpo
        celula_a14.font = fonte_corpo
        celula_b14.font = fonte_corpo
        celula_d14.font = fonte_corpo
        celula_a18.font = fonte_corpo
        celula_a19.font = fonte_corpo
        celula_a20.font = fonte_corpo
        celula_a21.font = fonte_corpo
        celula_a22.font = fonte_corpo
        celula_d18.font = fonte_corpo
        celula_d19.font = fonte_corpo
        celula_d20.font = fonte_corpo
        celula_d21.font = fonte_corpo
        celula_d22.font = fonte_corpo
        celula_a27.font = fonte_corpo
        celula_b27.font = fonte_corpo
        celula_c27.font = fonte_corpo
        celula_d27.font = fonte_corpo
        celula_e27.font = fonte_corpo
        celula_f27.font = fonte_corpo
        celula_f29.font = fonte_corpo
        celula_f30.font = fonte_corpo
        celula_e32.font = fonte_vermelha
        celula_a34.font = fonte_corpo
        celula_b4.font = fonte_corpo
        celula_b5.font = fonte_corpo
        celula_b6.font = fonte_corpo
        celula_a8.font = fonte_recibo
        celula_a10.font = fonte_titulo
        celula_a13.font = fonte_titulo
        celula_a17.font = fonte_titulo
        celula_a26.font = fonte_titulo
        celula_b10.font = fonte_titulo
        celula_b13.font = fonte_titulo
        celula_b26.font = fonte_titulo
        celula_b34.font = fonte_corpo
        celula_c26.font = fonte_titulo
        celula_d10.font = fonte_titulo
        celula_d13.font = fonte_titulo
        celula_d17.font = fonte_titulo
        celula_d26.font = fonte_titulo
        celula_e26.font = fonte_titulo
        celula_e29.font = fonte_corpo
        celula_e30.font = fonte_corpo
        celula_e32.font = fonte_vermelha
        celula_f26.font = fonte_titulo
        
        planilha.save(f'CARTAS-FRETE/CARTA-FRETE {numero_recibo} ({container_nome_motorista.get()}).xlsx')
        
        texto_confirmar_emissao.configure(text='NOTA EMITIDA COM SUCESSO', text_color='green')
        texto_confirmar_emissao.place(x=870, y= 900)
        
        app_2.after(3000, texto_confirmar_emissao.place_forget)

        
    except TypeError as erro:
        texto_confirmar_emissao.configure(text=f'{erro}', text_color='red')
        
    
    
    
texto_confirmar_emissao = ctk.CTkLabel(app_2 ,text='')



botao_emitir_carta_frete = ctk.CTkButton(app_2, text='Emitir carta frete', font=('Arial', 16, 'bold'), height=50, width=200, command=emitir)
botao_emitir_carta_frete.pack(pady=20)


######################################################################### JANELA LOGIN #################################################################################################################
# definir aparencia 
ctk.set_appearance_mode(cor_sistema)

# definir janelas
app = ctk.CTk()


# atribuir medidas para a tela 
largura_app = 300
altura_app = 300

# infirmações do tamanho da tela
largura_tela = app.winfo_screenwidth()
altura_tela = app.winfo_screenheight()

# centralizar o app no meio da tela 
x = int((largura_tela - largura_app) / 2)
y = int((altura_tela - altura_app) / 2)

# definir tamanho ja centralizado
app.geometry(f'{largura_app}x{altura_app}+{x}+{y}')

# titulo 
app.title('LOGIN')

#centralizar elementos
frame = ctk.CTkFrame(app, fg_color= 'transparent' )
frame.pack(expand=True)

# label da 'digite a senha'
texto_senha = ctk.CTkLabel(master=frame, text='Digite a senha:')
texto_senha.pack(pady=20, expand= True)

# adicionar container para a senha
container_senha = ctk.CTkEntry(master=frame, show='*', placeholder_text='Rodrivier Transportes')
container_senha.pack(pady= (0,20))

# botao para enviar a senha
def botao_enviar_senha(event= None):
    senha = container_senha.get().strip()
    if senha == 'yan12132701':
        app.destroy()
        app_2.mainloop()
        
        
    elif senha == '':
        feedback.configure(text= 'Insira a senha', text_color= 'yellow')
        
    else: 
        feedback.configure(text= 'Senha inrenavanreta', text_color= 'red')
                    
botao_login = ctk.CTkButton(master= frame, text='Entrar', command=botao_enviar_senha).pack()
app.bind('<Return>', botao_enviar_senha)
# feedback senha 
feedback = ctk.CTkLabel(master=frame, text= '', text_color='red')
feedback.pack()

img_icone = pilimage.open(caminho_recurso('imagens/icone.png'))
icone = ImageTk.PhotoImage(img_icone)
app_2.iconphoto(False, icone)

# iniciar tela
app.mainloop()  
        


